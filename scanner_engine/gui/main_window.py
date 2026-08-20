# --------------------------------------------------------------------------
# Copyright © 2025 Z-VulnScan Team. All Rights Reserved.
# 
# This software is proprietary and confidential. 
# Unauthorized copying, modification, distribution, or reverse engineering 
# of this file, via any medium, is strictly prohibited.
# --------------------------------------------------------------------------
import sys
import os
import csv
import re
import requests
import threading

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QProgressBar, QComboBox, QMessageBox,
    QSplitter, QTextEdit, QMenu,
    QGroupBox,
    QAbstractItemView,
    QSizePolicy,
    QCheckBox,
    QFileDialog,
    QSpinBox,
    QDialog,
    QStackedWidget,
    QScrollArea,
    QFrame,
    QSlider,
    QToolButton,
)
# [PySide6 핵심 기능]
from PySide6.QtCore import Qt, QTimer, QUrl, QSize
# [PySide6 그래픽 및 액션]
from PySide6.QtGui import QIcon, QColor, QBrush, QAction, QTextCursor, QKeySequence, QDesktopServices

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)


def resource_path(relative_path): 
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# 엔진 모듈 연동
from core.worker import ScanWorker              
from core.advanced_scanner import AdvancedScanner
from core.license_validator import LicenseValidator
from core.config import AppConfig
from output.pdf_report import PDFGenerator
from output.excel_report import ExcelGenerator
from output.text_report import TextReportGenerator
from utils.os_utils import OSUtils
from utils.secure_storage import SecureStorage
from utils.db_connector import DBConnector
from utils.logger import AppLogger
from utils.update_checker import check_for_updates
from gui.styles import STYLESHEET, LIGHT_STYLESHEET
from gui.help_texts import HELP_TEXTS
from gui.dialogs import LicenseDialog, PortSelectorDialog
from gui.dashboard_widgets import ScanConfigCard, MetricsRow, AssetsTableCard, RecentActivityCard, COLORS, set_theme as set_dashboard_theme, make_line_icon, HostnameSourceBadge, InfoCard, LabelWithHelp

class LicenseManager:
    #단일 바이너리 내에서 라이선스 등급에 따라 기능을 제어하는 매니저 클래스
    #- Standard: PDF만 가능 / 증적은 취약·부분만족 항목만 / 조치방안 없음
    #- Professional: Excel+PDF 가능 / 증적 전체 / 조치방안은 중요도 상·중 항목만
    #- Enterprise: 모든 기능(전문가 모드 포함) / 증적·조치방안 전체
    #
    TIER_STANDARD = "STANDARD"
    TIER_PROFESSIONAL = "PROFESSIONAL"
    TIER_ENTERPRISE = "ENTERPRISE"
    _TIER_CYCLE = [TIER_STANDARD, TIER_PROFESSIONAL, TIER_ENTERPRISE]

    def __init__(self):
        # [기본값 = Standard, 2026-08-19 정식 출시 대비 변경] 예전에는 license.dat에
        # 유효한 키가 없으면 본인 자체 사용 편의를 위해 Enterprise로 시작했으나, 그
        # 상태로 출시하면 라이선스 없이 받은 사람도 전체 기능을 무제한으로 쓸 수 있어
        # 상업적으로 말이 안 된다. 이제 키가 없거나 무효/만료면 가장 제한적인 Standard로
        # 폴백한다 - 유효한 키가 검증되는 즉시 ScannerApp.__init__에서 실제 등급으로 올라간다.
        self.current_tier = self.TIER_STANDARD
        # 실제 라이선스 키가 로드된 경우에만 채워짐 (없으면 무제한으로 취급)
        self.expiry_date = None

    def effective_tier(self):
        return self.current_tier

    def toggle_tier(self):
        """숨겨진 개발자 전용 동작(Ctrl+Shift+L, 메뉴/버튼에 노출되지 않음): 등급별
        제한이 실제로 어떻게 보이는지 확인하기 위해 STD -> PRO -> ENT 순으로 순환한다.
        메모리에서만 바뀌며 license.dat는 건드리지 않으므로, 재시작하면 원래 등급으로 돌아간다."""
        try:
            idx = self._TIER_CYCLE.index(self.current_tier)
        except ValueError:
            idx = -1
        self.current_tier = self._TIER_CYCLE[(idx + 1) % len(self._TIER_CYCLE)]
        return self.current_tier

    def get_window_title(self):
        #"""라이선스에 따른 윈도우 제목 반환"""
        base_title = f"Z-Vuln Scan {AppConfig.VERSION}"
        tier = self.effective_tier()
        if tier == self.TIER_STANDARD:
            return f"{base_title} Standard"
        elif tier == self.TIER_PROFESSIONAL:
            return f"{base_title} Professional"
        elif tier == self.TIER_ENTERPRISE:
            return f"{base_title} Enterprise"
        return base_title

    def can_export_excel(self):
        #"""Professional 등급 이상만 엑셀 내보내기 허용"""
        return self.effective_tier() in (self.TIER_PROFESSIONAL, self.TIER_ENTERPRISE)

    def can_use_expert_mode(self):
        #"""전문가 모드는 Enterprise 등급 전용"""
        return self.effective_tier() == self.TIER_ENTERPRISE

    def can_use_waiver(self):
        #"""[2026-08-06] 예외처리(Waiver)도 전문가 모드와 같은 Enterprise 등급 전용으로
        #통일 - 둘 다 "전문가용" 기능으로 묶어 설정 화면 한 곳에서만 접근하게 한다."""
        return self.effective_tier() == self.TIER_ENTERPRISE

    def evidence_level(self):
        """리포트에 상세 증적(raw_output)을 얼마나 보여줄지.
        'vuln_only': 취약/부분만족 항목만 / 'full': 전체 항목"""
        return "vuln_only" if self.effective_tier() == self.TIER_STANDARD else "full"

    def remediation_level(self):
        """리포트에 조치방안을 얼마나 보여줄지.
        'none': 미제공 / 'partial': 중요도 상·중 항목만 / 'full': 전체"""
        tier = self.effective_tier()
        if tier == self.TIER_STANDARD:
            return "none"
        elif tier == self.TIER_PROFESSIONAL:
            return "partial"
        return "full"  # ENTERPRISE
    
class ScannerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DBConnector()
        self.worker = None
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_timer)
        self.elapsed_seconds = 0
        
        # 스캔 결과 임시 저장소 (UI 렉 방지용 버퍼)
        self.scan_result_buffer = []

        # [P0: 자산 매핑] import_asset_list()로 가져온 {ip: hostname} 맵. hostname 폴백과
        # 스캔 종료 후 선언/실제 자산 diff에 쓰인다. 가져오기 전엔 빈 dict.
        self.imported_asset_map = {}

        # [Phase 2] 이번 스캔 결과를 Excel/PDF/TXT 중 하나로 한 번이라도 출력했는지 여부.
        # False인 상태에서 DB를 초기화하려 하면 경고를 띄운다 (출력 전 데이터 유실 방지).
        self.report_exported = True

        self.initUI()
        self.license_mgr = LicenseManager()
        saved_key = LicenseValidator.load_license()
        if saved_key:
            is_valid, tier, expiry_date = LicenseValidator.validate_key(saved_key)
            if is_valid:
                # 유효한 키가 있으면 그 등급/만료일로 올리거나 갱신 (기본값은 Standard)
                self.license_mgr.current_tier = tier
                self.license_mgr.expiry_date = expiry_date
                print(f"[System] Valid License Found: {tier} (expires {expiry_date})")
            else:
                print("[System] Saved license key invalid/expired - using default Standard mode")
        # [버그 수정] 예전엔 이 호출이 위 키 로딩보다 먼저 있어서, 실제 키가 있어도
        # 창 제목이 항상 기본값(Standard) 기준으로만 표시되던 문제가 있었음 - 반드시
        # current_tier가 최종 확정된 뒤에 호출해야 한다.
        self.update_ui_by_license()
        self.load_saved_data()
        self._check_for_updates()
        self._check_ruleset_update()

    # -- UI 세팅 --
    def initUI(self):
        self.setWindowTitle('Z-Vuln Scan')
        self.resize(1100, 750)
        self.setWindowIcon(QIcon(resource_path('app_icon.ico')))

        # [테마] 전체 스타일시트 (Deep Dark 모드 + 가독성 개선) - [Phase 3] 설정에서 라이트 테마 선택 가능
        from utils.app_settings import load_settings as _load_app_settings
        self._app_settings = _load_app_settings()
        theme = self._app_settings.get("theme", "light")
        self.setStyleSheet(LIGHT_STYLESHEET if theme == "light" else STYLESHEET)
        # [다크 테마 대응] 대시보드 카드(dashboard_widgets.py)는 COLORS를 직접 참조하므로,
        # 아래에서 위젯을 만들기 "전에" 반드시 먼저 호출해 카드들도 같은 테마를 쓰게 한다.
        set_dashboard_theme(theme)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        root_layout = QVBoxLayout(central_widget)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        root_layout.addWidget(self._build_topbar())

        body = QWidget()
        outer_layout = QHBoxLayout(body)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        outer_layout.addWidget(self._build_sidebar())

        self.content_stack = QStackedWidget()
        self.content_stack.addWidget(self._build_dashboard_page())  # index 0: Dashboard
        self.content_stack.addWidget(self._build_scan_page())       # index 1: Scan
        self.content_stack.addWidget(self._build_assets_page())     # index 2: Assets
        self.content_stack.addWidget(self._build_reports_page())    # index 3: Reports
        outer_layout.addWidget(self.content_stack, 1)

        root_layout.addWidget(body, 1)

        # [로그 패널] 도킹/플로팅 없이 메인 창에 고정 임베드. 기본은 숨김이며,
        # 설정(Settings > 테마·UI) 체크박스로 표시 여부를 켤 수 있다.
        self.log_panel = self._build_log_panel()
        self.log_panel.setVisible(self._app_settings.get("show_log_panel", False))
        root_layout.addWidget(self.log_panel)

        # [상태바] QMainWindow 기본 상태바 대신 central layout 하단에 직접 배치
        self.setStatusBar(None)
        status_bar_widget = QWidget()
        status_bar_widget.setStyleSheet(
            f"background-color: {COLORS['surface_1']}; color: {COLORS['text_secondary']}; border-top: 1px solid {COLORS['border']};"
        )
        status_layout = QHBoxLayout(status_bar_widget)
        status_layout.setContentsMargins(15, 5, 15, 5)

        self.time_label = QLabel("Ready")
        self.time_label.setStyleSheet(f"font-weight: bold; font-size: 9pt; color: {COLORS['text']};")

        self.pbar = QProgressBar()
        self.pbar.setFixedHeight(14)
        self.pbar.setTextVisible(False)
        self.pbar.setStyleSheet(
            f"QProgressBar {{ background: {COLORS['surface_2']}; border: 1px solid {COLORS['border']}; border-radius: 3px; }}"
            f"QProgressBar::chunk {{ background: {COLORS['accent']}; }}"
        )

        status_layout.addWidget(self.time_label)
        status_layout.addStretch()
        status_layout.addWidget(self.pbar, 1)

        root_layout.addWidget(status_bar_widget)

        # [숨겨진 개발자 전용 동작] 메뉴/버튼 어디에도 노출하지 않고, 단축키로만 접근한다.
        # 등급별 제한(전문가 모드 잠금 등)이 실제로 어떻게 보이는지 확인하기 위한 용도.
        self.action_license_switch = QAction("Cycle License Tier (Dev Only)", self)
        self.action_license_switch.setShortcut(QKeySequence("Ctrl+Shift+L"))
        self.action_license_switch.triggered.connect(self.demo_toggle_license)
        self.addAction(self.action_license_switch)

    # ------------------------------------------------------------------
    # [Phase 3: UI 재구성] 좌측 아이콘 사이드바
    # ------------------------------------------------------------------
    def _build_sidebar(self):
        # [UI/UX 개선 - "신뢰할 수 있는 작업대"] 아이콘만 있던 80px 좁은 사이드바를
        # 라벨이 살아있는 208px 리스트형으로 교체 - 국산 보안 콘솔/Qualys 계열이
        # 공유하는 사이드바 문법(아이콘+텍스트, 전체 폭 행, 좌측 액센트 스트라이프로
        # 활성 탭 표시)에 맞춘다. 목업(zvulnscan_design_directions) "방향 C" 기준.
        sidebar = QWidget()
        sidebar.setFixedWidth(208)
        sidebar.setStyleSheet(f"background-color: {COLORS['surface_2']}; border-right: 1px solid {COLORS['border']};")
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(14, 16, 14, 16)
        layout.setSpacing(2)

        brand = QLabel("Z-VulnScan")
        brand.setStyleSheet(f"font-size: 13pt; font-weight: 800; color: {COLORS['text']}; margin: 2px 6px 14px;")
        layout.addWidget(brand)

        def make_nav_group_label(text):
            lbl = QLabel(text.upper())
            lbl.setStyleSheet(
                f"font-size: 8pt; font-weight: 700; letter-spacing: 1px; "
                f"color: {COLORS['text_muted']}; margin: 10px 6px 4px;"
            )
            layout.addWidget(lbl)

        def make_nav_button(label_text, tooltip, icon_kind, checkable=True):
            btn = QToolButton()
            btn.setText(f" {label_text}")
            btn.setToolTip(tooltip)
            btn.setCheckable(checkable)
            btn.setFixedHeight(38)
            btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
            btn.setIconSize(QSize(17, 17))
            btn.setIcon(make_line_icon(icon_kind, COLORS['text_secondary']))
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(f"""
                QToolButton {{
                    font-size: 10pt; font-weight: 600; text-align: left;
                    border-radius: 8px; border-left: 3px solid transparent;
                    color: {COLORS['text_secondary']}; background: transparent;
                    padding-left: 8px;
                }}
                QToolButton:hover {{ background: {COLORS['accent_bg']}; color: {COLORS['text']}; }}
                QToolButton:checked {{
                    background: {COLORS['accent_bg']}; color: {COLORS['accent']};
                    border-left: 3px solid {COLORS['accent']};
                }}
            """)
            layout.addWidget(btn)
            # [활성 탭] 체크될 때 아이콘도 액센트 색으로 다시 그린다 - QSS는 QIcon
            # 픽셀 자체의 색은 못 바꾸므로 코드에서 직접 교체한다.
            if checkable:
                def _restyle_icon(checked):
                    btn.setIcon(make_line_icon(icon_kind, COLORS['accent'] if checked else COLORS['text_secondary']))
                btn.toggled.connect(_restyle_icon)
            return btn

        make_nav_group_label("진단")

        self.nav_dashboard = make_nav_button("대시보드", "Dashboard", "grid")
        self.nav_dashboard.setChecked(True)
        self.nav_dashboard.clicked.connect(lambda: self._on_nav_clicked("dashboard"))

        self.nav_scan = make_nav_button("스캔", HELP_TEXTS["discovery"]["title"] + " / " + HELP_TEXTS["audit"]["title"], "search")
        self.nav_scan.clicked.connect(lambda: self._on_nav_clicked("scan"))

        self.nav_assets = make_nav_button("자산", "Assets", "folder")
        self.nav_assets.clicked.connect(lambda: self._on_nav_clicked("assets"))

        self.nav_reports = make_nav_button("리포트", HELP_TEXTS["report"]["tooltip"], "doc")
        self.nav_reports.clicked.connect(lambda: self._on_nav_clicked("reports"))

        self.nav_crosscheck = make_nav_button("교차검증", "스크립트 TXT 결과를 오프라인으로 재판정/대조합니다 (DB 미사용).", "compare", checkable=False)
        self.nav_crosscheck.clicked.connect(self.open_cross_check)

        # [기능 비활성화] PC 진단 도구: 컨설턴트 스크립트 로직/방식을 아무리 따라가도 컨설턴트가
        # 제공하는 엑셀 결과물과 호환성을 맞출 수 없다는 사용자 판단으로 비활성화(코드는 보존,
        # 삭제하지 않음 - pc_toolkit.py/pc_toolkit_dialog.py/import_pc_results()는 그대로 둠).
        # 기존 WinRM 기반 PC 점검 경로(worker.py)는 이 결정과 무관하게 그대로 유지된다.
        self.nav_pc_toolkit = make_nav_button("PC 진단", "WinRM이 안 되는 PC용 로컬 진단 스크립트를 생성하고 결과를 가져옵니다.", "folder", checkable=False)
        self.nav_pc_toolkit.clicked.connect(self.open_pc_toolkit)
        self.nav_pc_toolkit.setVisible(False)

        layout.addStretch()

        make_nav_group_label("지원")

        self.nav_help = make_nav_button("도움말", HELP_TEXTS["settings"]["tooltip"], "help", checkable=False)
        self.nav_help.clicked.connect(self.open_help)

        self.nav_settings = make_nav_button("설정", HELP_TEXTS["settings"]["tooltip"], "settings", checkable=False)
        self.nav_settings.clicked.connect(self.open_settings)

        self._nav_buttons = {"dashboard": self.nav_dashboard, "scan": self.nav_scan,
                              "assets": self.nav_assets, "reports": self.nav_reports}
        return sidebar

    def _on_nav_clicked(self, key):
        for name, btn in self._nav_buttons.items():
            btn.setChecked(name == key)

        index_map = {"dashboard": 0, "scan": 1, "assets": 2, "reports": 3}
        self.content_stack.setCurrentIndex(index_map[key])

        crumb_labels = {"dashboard": "대시보드", "scan": "스캔", "assets": "자산", "reports": "리포트"}
        if hasattr(self, 'lbl_breadcrumb'):
            self.lbl_breadcrumb.setText(crumb_labels.get(key, ""))

    # ------------------------------------------------------------------
    # [UI/UX 개선 - "신뢰할 수 있는 작업대"] 상단바 - 목업의 breadcrumb 영역만 반영.
    # 브랜드는 이미 사이드바에 있어 중복 노출하지 않는다.
    # ------------------------------------------------------------------
    def _build_topbar(self):
        bar = QWidget()
        bar.setFixedHeight(46)
        bar.setStyleSheet(f"background-color: {COLORS['surface_2']}; border-bottom: 1px solid {COLORS['border']};")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(24, 0, 24, 0)

        self.lbl_breadcrumb = QLabel("대시보드")
        self.lbl_breadcrumb.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 10pt; font-weight: 600;")
        layout.addWidget(self.lbl_breadcrumb)
        layout.addStretch()

        # [디자인 목업 일치] 목업의 op-chip(우측 알약형 칩) - 담당자 이름(스캔 페이지
        # Operator 입력)과 라이선스 등급을 보여준다. 상단바 생성 시점(initUI() 초반)엔
        # license_mgr도 operator_input도 아직 없으므로 일단 제품명으로 채워두고,
        # 실제 값은 refresh_topbar_chip()이 둘 다 준비된 뒤 채운다.
        self.topbar_chip = QLabel("Z-VulnScan Professional Edition")
        self.topbar_chip.setStyleSheet(f"""
            color: {COLORS['text_secondary']}; font-size: 9pt; font-weight: 600;
            background: {COLORS['muted_bg']}; border-radius: 10px; padding: 4px 12px;
        """)
        layout.addWidget(self.topbar_chip)

        return bar

    def refresh_topbar_chip(self):
        """[UI/UX 개선] 상단바 칩을 '{담당자} · {라이선스 등급}'으로 채운다. 담당자를
        아직 안 적었으면 등급만, license_mgr가 아직 없으면(초기화 극초반) 건드리지
        않고 제품명 기본값을 그대로 둔다."""
        if not hasattr(self, 'topbar_chip') or not hasattr(self, 'license_mgr'):
            return
        tier_label = self.license_mgr.effective_tier().capitalize()
        operator = self.operator_input.text().strip() if hasattr(self, 'operator_input') else ""
        self.topbar_chip.setText(f"{operator} · {tier_label}" if operator else tier_label)

    def _wrap_scrollable_page(self, widgets, trailing_stretch=False):
        """대시보드/스캔 페이지 공통 - 세로 스크롤만 허용하는 카드 컨테이너.
        widgets: QWidget 또는 (QWidget, stretch) 튜플의 리스트."""
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        outer.addWidget(scroll)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(18, 14, 18, 16)
        layout.setSpacing(12)
        scroll.setWidget(content)

        for w in widgets:
            if isinstance(w, tuple):
                widget, stretch = w
                layout.addWidget(widget, stretch)
            else:
                layout.addWidget(w)

        if trailing_stretch:
            layout.addStretch()

        return page

    # ------------------------------------------------------------------
    # [Phase 3: UI 재구성] 대시보드 페이지 (요약 지표 + 결과 테이블 - 모니터링/조회 전용)
    # ------------------------------------------------------------------
    def _build_dashboard_page(self):
        # [UI/UX 개선 - 탭 역할 분리] 대시보드는 한눈에 보는 현황(지표 카드)만
        # 담당한다. 조치가 필요한 진단 결과표(Waiver/Expert/수동입력 버튼 포함)는
        # "무엇을 조치할지 결정하는" 화면이라 자산 탭이 더 자연스러운 위치라
        # _build_assets_page()로 옮겼다 - 여기선 더 이상 만들지 않는다.
        header = self._build_page_header("Dashboard")
        self.recent_activity_card = RecentActivityCard()
        self.recent_activity_card.session_clicked.connect(self.drill_down_session)
        self.recent_activity_card.period_changed.connect(lambda _days: self.refresh_recent_activity())
        # [버그 수정] trailing_stretch 없이는 스크롤 영역의 남는 세로 공간이 마지막
        # 위젯에 몰려서 내용이 비정상적으로 커지는 문제가 있었다 - 남는 공간을
        # 흡수할 stretch를 맨 끝에 명시적으로 둔다.
        return self._wrap_scrollable_page(
            [header, self._build_metrics_row(), self._build_unresolved_warning(), self.recent_activity_card],
            trailing_stretch=True,
        )

    def _build_unresolved_warning(self):
        """[커버리지 추적] 발견됐지만 KISA 판정이 하나도 없는 자산이 있으면 보여주고
        경고한다(감사 완결성 - "이 자산은 아무도 안 봤다"가 리포트 출력 시점까지
        조용히 묻히는 걸 막기 위함). 미해결 자산이 0건이면 숨긴다."""
        self.lbl_unresolved_warning = QLabel("")
        self.lbl_unresolved_warning.setWordWrap(True)
        # [버그 수정] QLabel 기본 수직 정책(Preferred)이 남는 레이아웃 공간을 그대로
        # 흡수해버려 한 줄짜리 텍스트가 화면 절반 높이로 늘어나는 문제가 있었다 -
        # 내용 높이(sizeHint)까지만 차지하도록 명시적으로 고정한다.
        self.lbl_unresolved_warning.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        self.lbl_unresolved_warning.setStyleSheet(
            f"background: {COLORS['danger_bg']}; color: {COLORS['danger_text']}; "
            f"border: 1px solid {COLORS['danger_text']}; border-radius: 8px; padding: 10px 14px;"
        )
        self.lbl_unresolved_warning.setVisible(False)
        return self.lbl_unresolved_warning

    def refresh_unresolved_warning(self):
        if not hasattr(self, 'lbl_unresolved_warning'):
            return
        try:
            unresolved = self.db.get_unresolved_assets()
        except Exception as e:
            AppLogger.log_error("[UI] Refresh Unresolved Warning Failed", e)
            return

        if not unresolved:
            self.lbl_unresolved_warning.setVisible(False)
            return

        shown = ", ".join((h or ip) for ip, h in unresolved[:10])
        more = f" 외 {len(unresolved) - 10}건" if len(unresolved) > 10 else ""
        self.lbl_unresolved_warning.setText(
            f"⚠ 미해결 자산 {len(unresolved)}건 — 발견됐지만 KISA 판정이 아직 하나도 없습니다: {shown}{more}"
        )
        self.lbl_unresolved_warning.setVisible(True)

    # ------------------------------------------------------------------
    # [Phase 3: UI 재구성] 스캔 페이지 (Scan configuration - 입력/실행 전용)
    # ------------------------------------------------------------------
    def _build_scan_page(self):
        header = self._build_page_header("Scan", show_operator=True)
        return self._wrap_scrollable_page([header, self._build_scan_config_card()], trailing_stretch=True)

    def _build_page_header(self, title_text, show_operator=False):
        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(8)

        accent_bar = QFrame()
        accent_bar.setFixedSize(4, 22)
        accent_bar.setStyleSheet(f"background-color: {COLORS['accent']}; border-radius: 2px;")
        header_layout.addWidget(accent_bar)

        self.title_label = QLabel(title_text)
        self.title_label.setStyleSheet(f"color: {COLORS['text']}; font-size: 16pt; font-weight: bold;")
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()

        if show_operator:
            # [Phase 2] 운영자 태깅 입력 - 스캔을 실행하는 화면에 배치
            header_layout.addWidget(QLabel("Operator:"))
            self.operator_input = QLineEdit()
            self.operator_input.setPlaceholderText("담당자")
            self.operator_input.setToolTip("이 스캔을 수행하는 담당자 이름입니다. 진단 결과에 함께 기록되어 감사 추적에 사용됩니다.")
            self.operator_input.setFixedWidth(160)
            self.operator_input.textChanged.connect(self.refresh_topbar_chip)
            header_layout.addWidget(self.operator_input)

        return header

    def _build_scan_config_card(self):
        """[Phase 3: UI 재구성] 사용자가 보내준 라이트 테마 설계(dashboard_widgets.ScanConfigCard)를
        그대로 쓰고, 기존 코드가 참조하는 속성명으로 alias만 걸어준다."""
        self.scan_config = ScanConfigCard()

        # 위젯 alias - main_window.py 나머지 코드는 예전 이름으로 그대로 접근
        self.ip_input = self.scan_config.target_input
        self.port_mode_combo = self.scan_config.scan_mode_combo
        self.port_input = self.scan_config.port_input
        self.port_select_btn = self.scan_config.port_select_btn
        self.user_input = self.scan_config.user_input
        self.pw_input = self.scan_config.pw_input
        self.db_cred_diff_check = self.scan_config.db_cred_diff_check
        self.db_user_input = self.scan_config.db_user_input
        self.db_pw_input = self.scan_config.db_pw_input
        self.oracle_service_input = self.scan_config.oracle_service_input
        self.slider_max_workers = self.scan_config.concurrency_slider
        self.spin_max_workers = self.scan_config.concurrency_slider  # 기존 코드 호환용 alias
        self.chk_ot_mode = self.scan_config.ot_mode_check
        self.chk_demo_mode = self.scan_config.demo_mode_check
        self.btn_scan = self.scan_config.discovery_btn
        self.btn_audit = self.scan_config.audit_btn
        self.btn_stop = self.scan_config.stop_btn

        default_user = getattr(self, '_app_settings', {}).get("default_username", "")
        if default_user:
            self.user_input.setText(default_user)

        self.ip_input.setToolTip(HELP_TEXTS["discovery"]["tooltip"])
        self.chk_ot_mode.setToolTip(HELP_TEXTS["ot_demo_mode"]["ot_tooltip"])
        self.chk_demo_mode.setToolTip(HELP_TEXTS["ot_demo_mode"]["demo_tooltip"])
        self.btn_scan.setToolTip(HELP_TEXTS["discovery"]["tooltip"])
        self.btn_audit.setToolTip(HELP_TEXTS["audit"]["tooltip"])

        self.port_mode_combo.currentIndexChanged.connect(self.toggle_port_input)
        self.port_select_btn.clicked.connect(self.open_port_selector)
        self.chk_ot_mode.toggled.connect(self._on_ot_mode_toggled)
        self.btn_scan.clicked.connect(self.start_network_scan)
        self.btn_audit.clicked.connect(self.start_audit)
        self.btn_stop.clicked.connect(self.stop_scan)

        return self.scan_config

    def _build_metrics_row(self):
        self.metrics_row = MetricsRow()
        # refresh_dashboard_metrics()가 예전처럼 metric_labels로 접근할 수 있게 alias
        self.metric_labels = {k: card.value_label for k, card in self.metrics_row.cards.items()}
        return self.metrics_row

    def _build_results_card(self):
        self.assets_table_card = AssetsTableCard()

        # [UI/UX 개선 - 2026-08-06] Waiver/Expert는 "전문가용" 기능이라 Settings >
        # 룰셋/전문가 탭에도 같은 기능이 있지만(중복은 의도적으로 유지), 실제 조치
        # 워크플로우와 붙어있는 자산 탭이 더 접근성이 좋아 여기 버튼도 유지한다.
        # 예전엔 라이선스가 부족해도 버튼이 항상 보이고 클릭하면 차단 메시지만
        # 떴는데, 이제는 Enterprise가 아니면 버튼 자체를 숨긴다(update_ui_by_license()
        # 가 갱신 - initUI() 시점엔 license_mgr가 아직 없어 일단 숨김으로 시작).
        self.btn_waiver = QPushButton("Waiver")
        self.btn_waiver.setToolTip(HELP_TEXTS["waiver"]["tooltip"])
        self.btn_waiver.clicked.connect(self.open_waiver_manager)
        self.btn_waiver.setVisible(False)
        self.assets_table_card.actions_layout.addWidget(self.btn_waiver)
        self.btn_expert = QPushButton("Expert")
        self.btn_expert.setToolTip(HELP_TEXTS["expert_mode"]["tooltip"])
        self.btn_expert.clicked.connect(self.open_expert_mode)
        self.btn_expert.setVisible(False)
        self.assets_table_card.actions_layout.addWidget(self.btn_expert)
        btn_manual_audit = QPushButton("수동 진단 입력")
        btn_manual_audit.setToolTip("원격 접속이 불가능한 자산을 육안으로 점검한 결과를 항목별로 직접 입력합니다.")
        btn_manual_audit.clicked.connect(self.open_manual_audit)
        self.assets_table_card.actions_layout.addWidget(btn_manual_audit)
        btn_refresh = QPushButton("Refresh")
        btn_refresh.clicked.connect(self.refresh_dashboard)
        self.assets_table_card.actions_layout.addWidget(btn_refresh)

        return self.assets_table_card

    # ------------------------------------------------------------------
    # [Phase 3: UI 재구성] Assets 페이지 (기존 자산 테이블 + 컨텍스트 메뉴 유지)
    # ------------------------------------------------------------------
    def _build_assets_page(self):
        page = QWidget()
        left_layout = QVBoxLayout(page)
        left_layout.setContentsMargins(28, 28, 28, 28)

        list_header = QHBoxLayout()
        lbl_list = QLabel("Asset List")
        lbl_list.setStyleSheet(f"font-weight: bold; color: {COLORS['text']}; font-size: 12pt;")

        self.btn_clear_assets = QPushButton("Clear List")
        self.btn_clear_assets.setFixedSize(90, 40)
        self.btn_clear_assets.setStyleSheet(f"""
            QPushButton {{ background: {COLORS['surface_1']}; color: {COLORS['text_secondary']}; border: 1px solid {COLORS['border']}; border-radius: 4px; }}
            QPushButton:hover {{ background: {COLORS['danger_bg']}; color: {COLORS['danger_text']}; border-color: {COLORS['danger_text']}; }}
        """)
        self.btn_clear_assets.clicked.connect(self.clear_asset_table)

        # [복구] Phase 3 사이드바 재구성 때 DB Manager로 가는 버튼이 누락되어
        # UI에서 열 방법이 없어졌던 문제 - Assets 페이지 헤더에 다시 연결
        self.btn_open_db_manager = QPushButton("DB Manager")
        self.btn_open_db_manager.setToolTip("자산 정보(호스트명/구역 태그/메모 등)를 직접 조회·수정·삭제합니다.")
        self.btn_open_db_manager.clicked.connect(self.open_db_manager)

        # [자산평가] Excel 리포트의 점검대상/호스트별 시트에 쓰일 용도/부서/담당자/C-I-A
        # 등급 입력 - 스캔이 채우는 값이 아니라 사람이 입력하는 값이라 DB Manager와
        # 별개 다이얼로그로 둔다.
        self.btn_open_asset_assessment = QPushButton("자산평가")
        self.btn_open_asset_assessment.setToolTip("Excel 리포트에 들어갈 자산 용도/부서/담당자/C·I·A 등급을 입력합니다.")
        self.btn_open_asset_assessment.clicked.connect(self.open_asset_assessment)

        # [자산 export - 2026-08-06 신규] 자산 목록 전체를 CSV/Excel로 내보낸다 -
        # 고객사 자산대장 갱신/백업용. import(정보자산목록 불러오기)의 반대 방향.
        self.btn_export_assets = QPushButton("내보내기")
        self.btn_export_assets.setToolTip("현재 등록된 자산 목록 전체(IP/hostname/구역태그/부서/담당자/설명 등)를 CSV 또는 Excel 파일로 저장합니다.")
        self.btn_export_assets.clicked.connect(self.export_asset_list)

        # [UI/UX 개선 - 2026-08-06] import 버튼을 스캔 탭에서 여기로 옮겼다 - 이제
        # import가 스캔 대상만 채우는 게 아니라 구역태그/부서/담당자/설명까지 DB에
        # 바로 반영하는 자산 등록 작업이라(import_asset_list() 참고) 자산 탭 쪽이
        # 더 맞는 자리라는 피드백 반영. export와 나란히 둬서 대칭도 맞춘다.
        self.btn_import_assets = QPushButton("가져오기")
        self.btn_import_assets.setToolTip(HELP_TEXTS["import_assets"]["tooltip"])
        self.btn_import_assets.clicked.connect(self.import_asset_list)

        list_header.addWidget(lbl_list)
        list_header.addStretch()
        list_header.addWidget(self.btn_open_db_manager)
        list_header.addWidget(self.btn_open_asset_assessment)
        list_header.addWidget(self.btn_import_assets)
        list_header.addWidget(self.btn_export_assets)
        list_header.addWidget(self.btn_clear_assets)
        left_layout.addLayout(list_header)

        self.asset_table = QTableWidget()
        self.asset_table.setColumnCount(7)
        self.asset_table.setHorizontalHeaderLabels(["IP Addr","hostname", "출처", "OS / Type","MAC Addr", "Description", "Vender"])
        self.asset_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.asset_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.asset_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.asset_table.verticalHeader().setVisible(False)
        self.asset_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.asset_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.asset_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.asset_table.setAlternatingRowColors(False)
        self.asset_table.setStyleSheet(f"""
            QTableWidget {{ background-color: {COLORS['surface_2']}; gridline-color: {COLORS['border']}; border: 1px solid {COLORS['border']}; border-radius: 4px; }}
            QHeaderView::section {{ background-color: {COLORS['surface_1']}; color: {COLORS['text_secondary']}; padding: 8px; border: none; border-bottom: 1px solid {COLORS['border']}; font-weight: bold; }}
            QTableWidget::item {{ padding: 5px; color: {COLORS['text']}; }}
            QTableWidget::item:selected {{ background-color: {COLORS['accent_bg']}; color: {COLORS['text']}; border-left: 2px solid {COLORS['accent']}; }}
            QTableWidget::item:hover {{ background-color: {COLORS['surface_1']}; }}
        """)
        self.asset_table.doubleClicked.connect(self.on_asset_double_click)
        self.asset_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.asset_table.customContextMenuRequested.connect(self.show_context_menu)

        left_layout.addWidget(self.asset_table)

        # [UI/UX 개선 - 탭 역할 분리] 진단 결과 열람 + 조치(Waiver/Expert/수동입력)는
        # 원래 대시보드에 있었으나, "이 자산에 대해 뭔가 조치한다"는 동선이 자산
        # 목록 바로 아래 있는 게 더 자연스러워 이 페이지로 옮겼다.
        results_header_row = QHBoxLayout()
        results_header = QLabel("진단 결과")
        results_header.setStyleSheet(f"font-weight: bold; color: {COLORS['text']}; font-size: 12pt; margin-top: 18px;")
        results_header_row.addWidget(results_header)
        results_header_row.addStretch()

        # [UI/UX 개선 - 최근 활동 드릴다운] 대시보드에서 특정 세션 행을 클릭하면
        # 여기가 그 세션 결과만으로 필터링되고, 이 배너로 알려준다. 기본은 숨김.
        self.lbl_session_filter = QLabel("")
        self.lbl_session_filter.setStyleSheet(
            f"color: {COLORS['accent']}; background: {COLORS['accent_bg']}; "
            f"border-radius: 8px; padding: 3px 10px; font-size: 10.5pt; margin-top: 18px;"
        )
        self.lbl_session_filter.setVisible(False)
        results_header_row.addWidget(self.lbl_session_filter)

        self.btn_clear_session_filter = QPushButton("전체 보기")
        self.btn_clear_session_filter.setFixedHeight(24)
        self.btn_clear_session_filter.setVisible(False)
        self.btn_clear_session_filter.clicked.connect(self._clear_session_filter)
        results_header_row.addWidget(self.btn_clear_session_filter)

        left_layout.addLayout(results_header_row)
        left_layout.addWidget(self._build_results_card())

        return page

    # ------------------------------------------------------------------
    # [UI/UX 개선 - 2026-08-06 신설] 리포트 탭 - 예전엔 사이드바 "리포트" 버튼이
    # 드롭다운 메뉴(PDF/Excel/TXT)만 여는 checkable=False 버튼이라 별도 화면이
    # 없었다. 커스터마이징(표지 제목/회사명/파일명) 입력이 생기면서 그 값을 둘 곳이
    # 필요해져 다른 탭들처럼 독립된 화면으로 분리했다("각 탭별로 완벽하게 분업"
    # 피드백 반영 - 대시보드=현황판, 스캔=설정, 자산=조치, 리포트=산출물 커스터마이징).
    # ------------------------------------------------------------------
    def _build_reports_page(self):
        header = self._build_page_header("Reports")

        card = InfoCard()
        outer = QVBoxLayout(card)
        outer.setContentsMargins(22, 18, 22, 18)
        outer.setSpacing(4)

        title = QLabel(f"<span style='color:{COLORS['accent']};'>&#9679;</span> 리포트 커스터마이징")
        title.setStyleSheet(f"font-size: 15px; font-weight: 600; color: {COLORS['text']}; border:none;")
        outer.addWidget(title)
        subtitle = QLabel("PDF/Excel 리포트에만 반영됩니다. TXT는 고객사 매크로 호환 형식이라 파일명/제목이 고정돼 있어 적용되지 않습니다.")
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px; border:none;")
        outer.addWidget(subtitle)
        outer.addSpacing(14)

        from utils.app_settings import load_settings as _load_app_settings
        saved = _load_app_settings()

        outer.addWidget(LabelWithHelp("회사명 / 브랜딩 문구", "표지에 제목 아래 작은 글씨로 들어갑니다. 비워두면 아예 표시하지 않습니다."))
        self.report_company_input = QLineEdit()
        self.report_company_input.setPlaceholderText("예: (주)OO공사 정보보안팀 (선택)")
        self.report_company_input.setText(saved.get("report_company_name", ""))
        outer.addWidget(self.report_company_input)
        outer.addSpacing(10)

        outer.addWidget(LabelWithHelp("문서 제목", "표지 맨 위 큰 제목입니다. 비워두면 기본 제목(주요정보통신기반시설 기술적 취약점 분석 평가 결과)을 씁니다."))
        self.report_title_input = QLineEdit()
        self.report_title_input.setPlaceholderText("주요정보통신기반시설 기술적 취약점 분석 평가 결과 (기본값)")
        self.report_title_input.setText(saved.get("report_title", ""))
        outer.addWidget(self.report_title_input)
        outer.addSpacing(10)

        outer.addWidget(LabelWithHelp("저장 파일명", "확장자·타임스탬프는 자동으로 붙습니다. 비워두면 기존처럼 자동 생성됩니다."))
        self.report_filename_input = QLineEdit()
        self.report_filename_input.setPlaceholderText("비워두면 자동 생성 (예: ZVulnScan_Report_날짜시간)")
        self.report_filename_input.setText(saved.get("report_filename", ""))
        outer.addWidget(self.report_filename_input)
        outer.addSpacing(16)

        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setStyleSheet(f"color: {COLORS['border']}; background-color: {COLORS['border']}; max-height: 1px;")
        outer.addWidget(divider)
        outer.addSpacing(14)

        format_label = QLabel("형식 선택")
        format_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px; font-weight: 600; border:none;")
        outer.addWidget(format_label)
        outer.addSpacing(6)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self.btn_report_pdf = QPushButton("PDF 생성")
        self.btn_report_pdf.setToolTip(HELP_TEXTS["report"]["tooltip"])
        self.btn_report_pdf.clicked.connect(self.generate_pdf)
        self.btn_report_excel = QPushButton("Excel 생성")
        self.btn_report_excel.setToolTip(HELP_TEXTS["report"]["tooltip"])
        self.btn_report_excel.clicked.connect(self.generate_excel)
        self.btn_report_txt = QPushButton("TXT 생성")
        self.btn_report_txt.setToolTip("호스트별 개별 파일, 고객사 매크로 호환 포맷입니다.")
        self.btn_report_txt.clicked.connect(self.generate_text_report)
        btn_row.addWidget(self.btn_report_pdf)
        btn_row.addWidget(self.btn_report_excel)
        btn_row.addWidget(self.btn_report_txt)
        btn_row.addStretch()
        outer.addLayout(btn_row)

        return self._wrap_scrollable_page([header, card], trailing_stretch=True)

    def drill_down_session(self, session_id):
        """[UI/UX 개선 - 드릴다운] 대시보드 '최근 활동'에서 세션 행을 클릭했을 때 호출.
        자산 탭으로 이동해 그 세션이 남긴 결과만 진단 결과표에 채운다."""
        self._on_nav_clicked("assets")
        try:
            findings = self.db.get_findings_by_session(session_id)
        except Exception as e:
            AppLogger.log_error("[UI] Drill Down Session Failed", e)
            findings = []

        self.assets_table_card.clear_rows()
        for ip, hostname, os_type, kisa_code, name, status, risk, waived in findings:
            host = hostname if hostname and hostname != "-" else ip
            self.assets_table_card.add_row(host, os_type or "", kisa_code or "", status or "", risk or "")

        self.lbl_session_filter.setText(f"세션 필터: {session_id} ({len(findings)}건)")
        self.lbl_session_filter.setVisible(True)
        self.btn_clear_session_filter.setVisible(True)

    def _clear_session_filter(self):
        self.lbl_session_filter.setVisible(False)
        self.btn_clear_session_filter.setVisible(False)
        self.refresh_findings_table()

    def _build_log_panel(self):
        """[Phase 3] 도킹/플로팅 창이 아니라 메인 창에 고정 임베드되는 로그 패널.
        기본적으로 숨겨져 있고, Settings > 테마·UI 탭의 체크박스로 표시 여부를 켠다."""
        panel = QWidget()
        panel.setFixedHeight(180)
        panel.setStyleSheet(
            f"background: {COLORS['surface_1']}; border-top: 1px solid {COLORS['border']};"
        )
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(12, 8, 12, 8)

        log_header = QHBoxLayout()
        log_title = QLabel("System Log")
        log_title.setStyleSheet(f"font-weight: bold; color: {COLORS['text']};")
        log_header.addWidget(log_title)
        log_header.addStretch()
        self.btn_clear_logs = QPushButton("Clear Log")
        self.btn_clear_logs.setFixedSize(90, 28)
        self.btn_clear_logs.clicked.connect(self.clear_logs)
        log_header.addWidget(self.btn_clear_logs)
        layout.addLayout(log_header)

        self.log_console = QTextEdit()
        self.log_console.setReadOnly(True)
        self.log_console.setStyleSheet(f"""
            QTextEdit {{
                background-color: {COLORS['surface_2']};
                color: {COLORS['text']};
                font-family: Consolas, 'Courier New', monospace;
                font-size: 9pt;
                border: 1px solid {COLORS['border']};
                border-radius: 4px;
                padding: 5px;
            }}
        """)
        layout.addWidget(self.log_console)

        return panel

    def refresh_dashboard_metrics(self):
        if not hasattr(self, 'metrics_row'):
            return
        try:
            metrics = self.db.get_dashboard_metrics()
            self.metrics_row.update_counts(metrics)
        except Exception as e:
            AppLogger.log_error("[UI] Refresh Dashboard Metrics Failed", e)
        try:
            trend = self.db.get_dashboard_trend()
            self.metrics_row.update_trend(trend)
        except Exception as e:
            AppLogger.log_error("[UI] Refresh Dashboard Trend Failed", e)
        self.refresh_unresolved_warning()
        self.refresh_recent_activity()

    def refresh_recent_activity(self):
        if not hasattr(self, 'recent_activity_card'):
            return
        try:
            sessions = self.db.get_recent_sessions(days=self.recent_activity_card.current_days())
        except Exception as e:
            AppLogger.log_error("[UI] Refresh Recent Activity Failed", e)
            sessions = []
        self.recent_activity_card.set_sessions(sessions)

    def refresh_findings_table(self):
        if not hasattr(self, 'assets_table_card'):
            return
        try:
            findings = self.db.get_all_latest_findings()
        except Exception:
            findings = []

        self.assets_table_card.clear_rows()
        for ip, hostname, os_type, kisa_code, name, status, risk, waived in findings:
            host = hostname if hostname and hostname != "-" else ip
            self.assets_table_card.add_row(host, os_type or "", kisa_code or "", status or "", risk or "")

    # --- 기능 메서드 ---
    # [Unified] 자산 추가 함수 통합 (중복 제거 및 기능 합침)
    def add_asset_to_table(self, ip, hostname, os_type, mac_addr, vendor, hostname_source=""):
        # 1. 중복 검사 (Cache 확인 -> 빠름)
        if not hasattr(self, 'scanned_ip_cache'):
            self.scanned_ip_cache = set()
            
        items = self.asset_table.findItems(ip, Qt.MatchExactly)
        row = -1
        
        if items:
            row = items[0].row() 
        else:
            row = self.asset_table.rowCount()
            self.asset_table.insertRow(row)
            self.scanned_ip_cache.add(ip)

        # 2. 설명 가져오기
        # [버그 수정] 발견된 호스트마다 새 DBConnector()를 만들면 매번 스키마
        # 마이그레이션(CREATE TABLE/ALTER TABLE 시도)이 재실행되고 커넥션도 새로
        # 열려서, /24 스캔(호스트 ~250개) 기준 GUI 스레드가 불필요하게 오래 막혔다.
        # __init__에서 이미 만들어 둔 self.db를 재사용한다.
        description_text = ""
        try:
            description_text = self.db.get_description(ip)
        except: pass

        # --- [핵심 수정] 강제 빈칸 처리 로직 ---
        # 입력된 값이 None이거나 "-" (대시)라면 무조건 빈 문자열("")로 바꿈
        def clean_text(text):
            if text is None: return ""
            s = str(text).strip()
            if s == "-": return ""  # 범인 제거!
            return s

        hostname = clean_text(hostname)
        os_type  = clean_text(os_type)
        mac_addr = clean_text(mac_addr)
        vendor   = clean_text(vendor)
        # ------------------------------------

        # 3. 아이템 생성
        item_ip = QTableWidgetItem(ip)
        item_host = QTableWidgetItem(hostname)
        item_os = QTableWidgetItem(os_type)
        item_mac = QTableWidgetItem(mac_addr)
        item_description = QTableWidgetItem(description_text if description_text else "")
        item_vendor = QTableWidgetItem(vendor)

        # 4. 스타일링 (색상 설정) - 라이트 테마 배경(흰색) 위에서도 읽히는 색상 사용
        item_ip.setForeground(QBrush(QColor(COLORS["text"])))
        item_host.setForeground(QBrush(QColor(COLORS["text_secondary"])))
        item_mac.setForeground(QBrush(QColor(COLORS["text_muted"])))
        item_vendor.setForeground(QBrush(QColor(COLORS["text_muted"])))
        item_description.setForeground(QBrush(QColor(COLORS["success_text"])))

        if "Linux" in os_type or "Ubuntu" in os_type:
            item_os.setForeground(QBrush(QColor("#B36B00")))
        elif "Windows" in os_type:
            item_os.setForeground(QBrush(QColor("#0B6FB3")))
        else:
            item_os.setForeground(QBrush(QColor(COLORS["text_muted"])))

        # 가운데 정렬
        for item in [item_ip, item_host, item_os, item_mac, item_description, item_vendor]:
            item.setTextAlignment(Qt.AlignCenter)

        # 5. 테이블 배치
        self.asset_table.setItem(row, 0, item_ip)
        self.asset_table.setItem(row, 1, item_host)
        # [UI/UX 개선 - hostname 출처 배지] 역DNS/매핑/실측/추정 중 어디서 hostname을
        # 얻었는지 - OT망 hostname 확보가 이 제품의 차별점이라 근거를 화면에서 직접 보여준다.
        self.asset_table.setCellWidget(row, 2, HostnameSourceBadge(hostname_source))
        self.asset_table.setItem(row, 3, item_os)
        self.asset_table.setItem(row, 4, item_mac)
        self.asset_table.setItem(row, 5, item_description)
        self.asset_table.setItem(row, 6, item_vendor)

    def clear_asset_table(self):
        if self.asset_table.rowCount() == 0:
            return

        # [Phase 2] 아직 Excel/PDF/TXT로 한 번도 출력하지 않은 스캔 결과가 있으면 별도 경고
        if not self.report_exported:
            warn_reply = QMessageBox.warning(
                self, "리포트 미출력 경고",
                "이번 스캔 결과를 아직 Excel/PDF/TXT로 한 번도 출력하지 않았습니다.\n"
                "지금 초기화하면 해당 결과가 그대로 사라집니다.\n\n"
                "정말로 출력 없이 초기화를 계속하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if warn_reply != QMessageBox.Yes:
                return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            "모든 자산 데이터와 메모가 삭제됩니다.\n초기화하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            db = DBConnector()
            if db.delete_all_assets():
                self.asset_table.setRowCount(0)
                if hasattr(self, 'scanned_ip_cache'):
                    self.scanned_ip_cache.clear()
                self.report_exported = True
                self.log_message("[System] All assets cleared.")
                QMessageBox.information(self, "Cleared", "초기화되었습니다.")

    def clear_logs(self):
        self.log_console.clear()
        self.log_message("[UI] Log console cleared.")

    def on_asset_double_click(self):
        row = self.asset_table.currentRow()
        if row >= 0:
            ip = self.asset_table.item(row, 0).text()
            self.ip_input.setText(ip)
            self.log_message(f"[UI] Target Selected: {ip}")
            
            if ip in ["127.0.0.1", "localhost"]:
                self.user_input.setText("root")
                self.pw_input.setText("toor")
            elif ip == "0.0.0.0":
                self.user_input.setText("Administrator")
                self.pw_input.setText("password123")

    def update_timer(self):
        self.elapsed_seconds += 1
        
        def format_time(seconds):
            m, s = divmod(seconds, 60)
            h, m = divmod(m, 60)
            return f"{h:02d}:{m:02d}:{s:02d}"

        elapsed_str = format_time(self.elapsed_seconds)
        
        # 상태 메시지 단순화
        if self.pbar.value() >= 100:
            status_txt = f"Done (Duration: {elapsed_str})"
        else:
            status_txt = f"Running... [{elapsed_str}]"

        self.time_label.setText(status_txt)

    def update_progress(self, percent, count):
        self.pbar.setValue(percent)
        self.current_scan_count = count

    def log_message(self, msg):
        """[Fix] 로그 창 크래시 해결 및 최적화"""
        # 1. 텍스트 추가
        self.log_console.append(msg)
        
        # 2. 줄 수 제한 (1000줄)
        doc = self.log_console.document()
        if doc.blockCount() > 1000:
            cursor = self.log_console.textCursor()
            cursor.movePosition(cursor.Start)
            cursor.movePosition(cursor.Down, cursor.KeepAnchor, 100)
            cursor.removeSelectedText()
            
        # 3. [Fix] 올바른 MoveOperation 상수 사용
        self.log_console.moveCursor(QTextCursor.MoveOperation.End)
        self.log_console.ensureCursorVisible()

    def set_ui_busy(self, busy):
        # [Optimized] UI 상태 제어
        self.btn_scan.setDisabled(busy)
        self.btn_audit.setDisabled(busy)
        self.btn_stop.setEnabled(busy)
        
        self.nav_reports.setDisabled(busy)
        self.nav_assets.setDisabled(busy)
        # [스캔 중 화면 이동] Dashboard/Scan 탭은 스캔 도중에도 이동 가능하게 유지한다.
        # (입력 위젯들은 아래에서 별도로 비활성화되므로 설정을 바꿀 순 없고, 조회만 가능)

        self.ip_input.setDisabled(busy)
        self.port_mode_combo.setDisabled(busy)
        self.chk_ot_mode.setDisabled(busy)
        self.chk_demo_mode.setDisabled(busy)
        self.spin_max_workers.setDisabled(busy)
        self.user_input.setDisabled(busy)
        self.pw_input.setDisabled(busy)

        if busy:
            self.port_input.setDisabled(True)
            self.scan_result_buffer = []
            if hasattr(self, 'scanned_ip_cache'):
                self.scanned_ip_cache.clear()
            else:
                self.scanned_ip_cache = set()
            self.pbar.setValue(0)
            self.elapsed_seconds = 0
            self.time_label.setText("Initializing...")
        else:
            is_custom = (self.port_mode_combo.currentIndex() == 1)
            self.port_input.setEnabled(is_custom)
            self.timer.stop()
            self.pbar.setValue(100)

    def update_asset_table(self, ip, hostname, os_type, mac_addr, vendor, hostname_source=""):
        """[Optimized] 실시간 테이블 렌더링 대신 버퍼에 저장"""
        # 로그는 이미 실시간으로 뜨므로, 테이블 데이터는 모아둡니다.
        self.scan_result_buffer.append((ip, hostname, os_type, mac_addr, vendor, hostname_source))

    def _asset_declaration_diff_note(self):
        """[P0: 자산 매핑] 가져온 자산목록(선언)과 이번 스캔에서 실제 발견된 자산을
        비교해 누락(선언됐지만 응답 없음)/미문서화(발견됐지만 목록에 없음) 자산을
        알려준다. 감사 완결성 확인용 - 목록에 없는 미문서화 자산은 그 자체로
        KISA 감사에서 지적 대상이 될 수 있다. imported_asset_map이 비어있으면(이번
        세션에 목록을 안 가져왔으면) 비교 대상이 없으므로 건너뛴다."""
        if not self.imported_asset_map:
            return ""

        declared = set(self.imported_asset_map.keys())
        discovered = set(ip for ip, *_ in self.scan_result_buffer)
        missing = sorted(declared - discovered)
        undocumented = sorted(discovered - declared)

        if not missing and not undocumented:
            return "\n\n[자산목록 대조] 선언된 자산과 실제 발견 결과가 정확히 일치합니다."

        lines = ["\n\n[자산목록 대조]"]
        if missing:
            shown = ", ".join(missing[:10]) + (" ..." if len(missing) > 10 else "")
            lines.append(f"- 목록엔 있지만 응답 없음 ({len(missing)}건): {shown}")
        if undocumented:
            shown = ", ".join(undocumented[:10]) + (" ..." if len(undocumented) > 10 else "")
            lines.append(f"- 목록엔 없지만 발견됨 ({len(undocumented)}건, 미문서화 자산 가능성): {shown}")
        return "\n".join(lines)

    def scan_finished(self, msg):
        self.timer.stop()
        self.pbar.setValue(100)
        self.set_ui_busy(False)

        # [Phase 2] 실제로 뭔가 점검을 수행한 경우(Discovery/Audit 공통) 아직 리포트로
        # 출력되지 않은 새 결과가 생긴 것으로 표시한다. Audit이 캐시된 Discovery를 재사용하면
        # scan_result_buffer는 비어있을 수 있어도 새 진단(W-xx 등) 결과는 DB에 쌓이므로,
        # "대상이 아예 없었던" 경우만 제외하고 그 외에는 모두 False로 처리한다.
        if msg not in ("No Targets", "No Live Hosts", "Security Error"):
            self.report_exported = False

        if self.scan_result_buffer:
            count = len(self.scan_result_buffer)
            self.log_message(f"[System] Registering {count} assets to table...")
            
            self.asset_table.setSortingEnabled(False) 
            self.asset_table.setUpdatesEnabled(False) 
            
            for data in self.scan_result_buffer:
                ip, hostname, os_type, mac_addr, vendor, hostname_source = data
                self.add_asset_to_table(ip, hostname, os_type, mac_addr, vendor, hostname_source)
                
            self.asset_table.setUpdatesEnabled(True) 
            self.asset_table.setSortingEnabled(True) 
            self.log_message("[System] Table update completed.")
        
        diff_note = self._asset_declaration_diff_note()
        if diff_note:
            self.log_message("[System]" + diff_note.replace("\n\n", " ").replace("\n", " "))
        QMessageBox.information(self, "Finished", f"{msg}\n(Total Found: {len(self.scan_result_buffer)}){diff_note}")

        self.btn_scan.setEnabled(True)
        self.btn_audit.setEnabled(True)
        
        self.time_label.setText("Ready")
        self.time_label.setStyleSheet("font-weight: bold; font-size: 9pt; color: #ddd;")
        
        self.pbar.setValue(0) 

        self.log_message(f"[System] Scan Finished. (Total: {self.elapsed_seconds}s)")
        
        target_ip = self.ip_input.text().strip()
        user = self.user_input.text().strip()
        SecureStorage.delete_credential(target_ip, user)
        
        # [2단계 연동] 스캔 종료 시 위험 IP 미들웨어로 보고 (실제 진단 결과 기준 취약 건수)
        # [버그 수정] Discovery(NETWORK_SCAN)만 끝나도 보고되던 문제 - Audit(AUDIT_VULN)이
        # 완료됐을 때만 보고한다 (Discovery는 딥 인스펙션을 하지 않으므로 취약점 수 자체가 무의미함)
        is_audit_run = bool(self.worker) and getattr(self.worker, 'mode', None) == "AUDIT_VULN"
        if target_ip and is_audit_run:
            actual_vuln_count = self.db.get_vuln_count(target_ip)
            self.log_message(f"[System] 미들웨어에 취약점 진단 결과 전송 중... ({target_ip}, 취약 {actual_vuln_count}건)")
            threading.Thread(target=self.report_to_middleware, args=(target_ip, actual_vuln_count), daemon=True).start()
        
    def report_to_middleware(self, ip, vuln_count):
        url = "http://127.0.0.1:8089/api/v1/vuln_report"
        payload = {"target_ip": ip, "vuln_count": vuln_count}
        try:
            requests.post(url, json=payload, timeout=3)
        except:
            pass

    def start_network_scan(self):
        self.btn_scan.setEnabled(False) 
        self.btn_audit.setEnabled(False)
        
        ip = self.ip_input.text().strip()
        if not ip:
            QMessageBox.warning(self, "Input Error", "Target IP를 입력해주세요.")
            self.ip_input.setFocus()
            self.set_ui_busy(False)
            return

        mode_idx = self.port_mode_combo.currentIndex()
        target_ports = None 

        if mode_idx == 1: 
            p_str = self.port_input.text().strip()
            if not p_str:
                QMessageBox.warning(self, "Input Error", "Custom 포트를 입력해주세요.")
                self.set_ui_busy(False)
                return
            target_ports = AdvancedScanner.parse_ports(p_str)
            if not target_ports:
                QMessageBox.warning(self, "Error", "포트 형식이 올바르지 않습니다.")
                self.set_ui_busy(False)
                return
        elif mode_idx == 2: 
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Full Scan Warning")
            msg.setText("<b>전체 포트(1-65535) 스캔을 시작하시겠습니까?</b>")
            msg.setInformativeText("시간이 매우 오래 걸리며 네트워크 부하가 발생할 수 있습니다.")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg.setDefaultButton(QMessageBox.No)
            if msg.exec() == QMessageBox.No: 
                self.set_ui_busy(False)
                return
            target_ports = list(range(1, 65536))
        #self.asset_table.setRowCount(0)
        self.log_message("--- New Scan Started ---")
        
        # 중복 방지 캐시 초기화
        if hasattr(self, 'scanned_ip_cache'):
            self.scanned_ip_cache.clear()
        else:
            self.scanned_ip_cache = set()

        self.set_ui_busy(True)
        try:
            # None을 넘기면 Worker가 "Full Scan"으로 인식하게 됩니다.
            self.worker = ScanWorker("NETWORK_SCAN", ip, ports=target_ports, ot_mode=self.chk_ot_mode.isChecked(), demo_mode=self.chk_demo_mode.isChecked(), operator=self.operator_input.text().strip(), max_workers=self.spin_max_workers.value(), imported_hostname_map=self.imported_asset_map)
            self.connect_worker()
            self.worker.start()
        except Exception as e:
            self.log_message(f"[Error] Failed to start scan: {e}")
            self.set_ui_busy(False)

    def start_audit(self):
        ip = self.ip_input.text().strip()
        if not ip:
            QMessageBox.warning(self, "Error", "Target IP를 입력해주세요.")
            return
        if "/" in ip:
            QMessageBox.warning(self, "Notice", "Audit은 단일 IP만 지원합니다.")
            return

        is_sim = ip in ["127.0.0.1", "localhost", "0.0.0.0"]
        user = self.user_input.text().strip()
        pw = self.pw_input.text().strip()

        if not is_sim and (not user or not pw):
            QMessageBox.warning(self, "Auth Error", "SSH/WinRM 계정 정보가 필요합니다.")
            return

        if not is_sim:
            if not SecureStorage.save_credential(ip, user, pw):
                QMessageBox.critical(self, "Error", "자격증명 저장 실패")
                return

        # [DB 전용 계정] 체크했는데 DB user/pw 중 하나라도 비어있으면 실수로 빈 값이
        # 저장/사용되지 않도록 여기서 막는다. 체크 안 했으면 기존처럼 위 SSH/WinRM 계정을 그대로 쓴다.
        db_user = ""
        if self.db_cred_diff_check.isChecked():
            db_user = self.db_user_input.text().strip()
            db_pw = self.db_pw_input.text().strip()
            if not is_sim and (not db_user or not db_pw):
                QMessageBox.warning(self, "Auth Error", "DB 전용 계정을 사용하려면 DB User/Password를 모두 입력하세요.")
                return
            if not is_sim:
                if not SecureStorage.save_credential(ip, db_user, db_pw):
                    QMessageBox.critical(self, "Error", "DB 자격증명 저장 실패")
                    return

        oracle_service = self.oracle_service_input.text().strip()

        self.set_ui_busy(True)
        self.worker = ScanWorker("AUDIT_VULN", ip, user, db_user=db_user or None, ot_mode=self.chk_ot_mode.isChecked(), demo_mode=self.chk_demo_mode.isChecked(), operator=self.operator_input.text().strip(), max_workers=self.spin_max_workers.value(), imported_hostname_map=self.imported_asset_map, oracle_service_name=oracle_service or None)
        self.connect_worker()
        self.worker.start()

    def _current_report_customization(self):
        """[리포트 탭] 표지 제목/회사명/저장 파일명 입력값을 읽고, 다음 실행에서도
        이어쓸 수 있게 설정 파일에 저장한다. TXT는 고객사 매크로 호환 형식이라
        파일명 패턴이 고정돼야 해서 이 커스터마이징을 적용하지 않는다(PDF/Excel만)."""
        title = self.report_title_input.text().strip() if hasattr(self, 'report_title_input') else ""
        company = self.report_company_input.text().strip() if hasattr(self, 'report_company_input') else ""
        filename = self.report_filename_input.text().strip() if hasattr(self, 'report_filename_input') else ""
        try:
            from utils.app_settings import load_settings as _load_app_settings, save_settings as _save_app_settings
            settings = _load_app_settings()
            settings["report_title"] = title
            settings["report_company_name"] = company
            settings["report_filename"] = filename
            _save_app_settings(settings)
        except Exception as e:
            AppLogger.log_error("[UI] Save Report Customization Failed", e)
        return title, company, filename

    def generate_pdf(self):
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            title, company, filename = self._current_report_customization()
            generator = PDFGenerator(remediation_level=self.license_mgr.remediation_level(),
                                      report_title=title, company_name=company, custom_filename=filename)
            filepath = generator.generate()
            QApplication.restoreOverrideCursor()
            self.report_exported = True
            self.log_message(f"[Success] PDF Saved: {filepath}")
            if QMessageBox.question(self, "Success", "PDF를 여시겠습니까?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
                self.open_file_platform_safe(filepath)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(self, "Error", str(e))

    def generate_text_report(self):
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            # [버그 수정] Excel/PDF와 동일하게 라이선스 등급을 전달 - 예전엔 안 넘겨서
            # TXT로 내보내면 등급과 무관하게 증적/조치방안이 항상 전체 노출됐다.
            generator = TextReportGenerator(
                evidence_level=self.license_mgr.evidence_level(),
                remediation_level=self.license_mgr.remediation_level(),
            )
            # [매크로 호환] 호스트별로 별도 .txt 파일이 생성되므로 단일 파일이 아니라 목록으로 반환됨
            filepaths = generator.generate()
            QApplication.restoreOverrideCursor()
            self.report_exported = True
            out_dir = os.path.dirname(filepaths[0])
            self.log_message(f"[Success] TXT Report Saved: {len(filepaths)}개 파일 -> {out_dir}")
            if QMessageBox.question(self, "Success", f"호스트별 TXT 파일 {len(filepaths)}개가 생성되었습니다.\n저장 폴더를 여시겠습니까?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
                self.open_file_platform_safe(out_dir)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(self, "Error", str(e))

    def generate_excel(self):
        if not self.license_mgr.can_export_excel():
            QMessageBox.warning(self, "License Restricted", "이 기능은 Professional 라이선스 전용입니다.\n(데모: 툴바의 열쇠 버튼을 눌러보세요)")
            return
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            title, company, filename = self._current_report_customization()
            generator = ExcelGenerator(
                evidence_level=self.license_mgr.evidence_level(),
                remediation_level=self.license_mgr.remediation_level(),
                report_title=title, company_name=company, custom_filename=filename
            )
            filepath = generator.generate()
            QApplication.restoreOverrideCursor()
            self.report_exported = True
            self.log_message(f"[Success] Excel Saved: {filepath}")
            if QMessageBox.question(self, "Success", "Excel을 여시겠습니까?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
                self.open_file_platform_safe(filepath)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(self, "Error", str(e))

    def open_file_platform_safe(self, filepath):
        if not OSUtils.open_file(filepath):
            QMessageBox.warning(self, "Error", f"파일을 열 수 없습니다:\n{filepath}")

    def stop_scan(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop_flag = True
            self.btn_stop.setEnabled(False)
            self.log_message("[!!!] Stopping Scan...")

    def connect_worker(self):
        self.worker.log_signal.connect(self.log_message)
        self.worker.finish_signal.connect(self.scan_finished)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.started_signal.connect(self.on_scan_started)
        # [Optimized] 스캔 결과를 버퍼에 담는 함수 연결
        self.worker.asset_found_signal.connect(self.update_asset_table)

    def on_scan_started(self, count):
        self.log_message(f"[Info] Scanning Start. Targets: {count}")
        self.total_scan_count = count
        self.current_scan_count = 0
        self.timer.start(1000)

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.stop_flag = True
            self.worker.wait(2000)
        
        # 2. [추가됨] 종료 시 자격증명(ID/PW) 보안 삭제
        target_ip = self.ip_input.text().strip()
        user = self.user_input.text().strip()
        if target_ip and user:
            SecureStorage.delete_credential(target_ip, user)
            # 디버깅용 로그 (필요 시 주석 처리)
            print(f"[System] Credentials for {target_ip} wiped.")
        
        event.accept()

    def _on_ot_mode_toggled(self, checked):
        # [Phase 3] OT 모드를 켜면 동시처리량을 3으로 낮춰서 제안하되, 이미 더 낮게
        # 설정돼 있으면 건드리지 않는다 (사용자가 더 낮춰둔 값을 존중)
        if checked and self.spin_max_workers.value() > 3:
            self.spin_max_workers.setValue(3)

    def toggle_port_input(self, index):
        # 1. 인덱스가 1(Custom Range)일 때만 입력창 활성화
        is_custom = (index == 1)
        self.port_input.setEnabled(is_custom)
        self.port_select_btn.setEnabled(is_custom)

        # 2. 편의성: Custom 모드면 포트 입력창에 바로 포커스, 아니면 내용 지우기
        if is_custom:
            self.port_input.setFocus()
        else:
            self.port_input.clear()

    def open_port_selector(self):
        """[UI/UX 개선 - 2026-08-06] Custom 포트를 체크박스로 고를 수 있는 화면."""
        dlg = PortSelectorDialog(self.port_input.text(), self)
        if dlg.exec() == QDialog.Accepted:
            self.port_input.setText(dlg.get_ports_string())

    def load_saved_assets(self):
        # History 로드 시에는 Buffer가 아닌 즉시 등록 사용 (DB 내용이니까)
        db = DBConnector()
        assets = db.get_all_assets()
        
        self.asset_table.setRowCount(0)
        self.scanned_ip_cache = set() # 캐시 초기화
        
        # [버그 수정] get_all_assets()가 원래 hostname을 반환하지 않아 재시작/이력 로드
        # 때마다 Host 칸이 항상 빈칸이던 문제가 있었다 - 이제 hostname/hostname_source도
        # 함께 조회해 채운다(vendor는 여전히 이 쿼리로 조회되지 않아 빈 값).
        for ip, os_type, description, mac_addr, hostname, hostname_source in assets:
            display_os = os_type
            if mac_addr: display_os = f"{os_type} | {mac_addr}"
            self.add_asset_to_table(ip, hostname, display_os, mac_addr, "", hostname_source)

        if assets:
            self.log_message(f"[System] Loaded {len(assets)} assets from history.")

    def show_context_menu(self, pos):
        if not self.asset_table.selectionModel().selection().indexes(): return
        
        row = self.asset_table.currentRow()
        ip = self.asset_table.item(row, 0).text()
        # [버그 수정] 컬럼에 "출처" 배지가 새로 끼어들면서 OS/Type이 3번 컬럼으로 밀렸다 -
        # 기존엔 1번(hostname) 컬럼 텍스트를 os_type으로 잘못 읽어서 RDP/SSH 메뉴가
        # hostname 문자열에 "Windows"/"Linux"가 우연히 포함될 때만 뜨는 잠재 버그였다.
        os_type = self.asset_table.item(row, 3).text()

        menu = QMenu()
        menu.addAction(f"Ping Check ({ip})", lambda: OSUtils.open_ping_test(ip))
        menu.addSeparator()

        if "Windows" in os_type:
            menu.addAction("RDP Connect", lambda: OSUtils.open_rdp(ip))
        elif "Linux" in os_type:
            user = self.user_input.text().strip() or "root"
            menu.addAction("SSH Connect", lambda: OSUtils.open_ssh(ip, user))

        menu.addAction("Copy IP", lambda: QApplication.clipboard().setText(ip))
        menu.exec(self.asset_table.viewport().mapToGlobal(pos))
        
    def import_asset_list(self):
        """
        [Phase 1] 고객사 제공 정보자산목록(CSV/Excel)에서 IP를 읽어와 스캔 대상 입력란에 채운다.
        협의된 범위(파일)와 실제 스캔 대상이 1:1로 일치한다는 감사 증적이 되고,
        목록에 없는 IP는 애초에 후보에 들어가지 않는 "하드 제외" 효과도 겸한다.
        """
        filepath, _ = QFileDialog.getOpenFileName(
            self, "정보자산목록 불러오기", "",
            "자산목록 파일 (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls);;모든 파일 (*)"
        )
        if not filepath:
            return

        try:
            if filepath.lower().endswith((".xlsx", ".xls")):
                rows = self._read_asset_rows_excel(filepath)
            else:
                rows = self._read_asset_rows_csv(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"파일을 읽는 중 오류가 발생했습니다:\n{e}")
            return

        if not rows:
            QMessageBox.warning(self, "Notice", "파일에서 읽어들인 행이 없습니다.")
            return

        ip_pattern = re.compile(r'^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})$')

        def _looks_like_ip(cell):
            cell = str(cell).strip()
            m = ip_pattern.match(cell)
            if not m:
                return None
            if not all(0 <= int(g) <= 255 for g in m.groups()):
                return None
            return cell

        # 1) 헤더에서 "IP"가 들어간 컬럼, 그리고 자산명/호스트명/구역태그/부서/담당자/설명
        # 컬럼을 탐색한다. [P0: 자산 매핑] hostname 컬럼도 같이 읽으면 DNS/SNMP가 없는
        # OT 구간에서도 고객이 이미 문서로 갖고 있는 이름을 hostname 폴백 체인의 한
        # 단계로 쓸 수 있고, 스캔 후 "선언된 자산 vs 실제 발견된 자산" diff(누락/유령
        # 자산 탐지)도 가능해진다. [자산 import 강화 - 2026-08-06] 구역태그/부서/
        # 담당자/설명 컬럼도 인식해서 import 시점에 바로 DB에 반영한다(스캔을 굳이
        # 안 해도 정보자산목록 자체를 자산 등록/갱신 소스로 쓸 수 있게).
        header = rows[0]
        ip_col = None
        host_col = None
        zone_col = None
        dept_col = None
        owner_col = None
        desc_col = None
        HOST_HEADER_RE = re.compile(r'host|이름|명칭|자산명|장비명|호스트', re.IGNORECASE)
        ZONE_HEADER_RE = re.compile(r'구역|zone|영역|망분리|태그', re.IGNORECASE)
        DEPT_HEADER_RE = re.compile(r'부서|department|팀', re.IGNORECASE)
        OWNER_HEADER_RE = re.compile(r'담당자|담당|owner|manager', re.IGNORECASE)
        DESC_HEADER_RE = re.compile(r'설명|비고|description|note|remark', re.IGNORECASE)
        for idx, col in enumerate(header):
            col_str = str(col) if col else ""
            if ip_col is None and re.search(r'ip', col_str, re.IGNORECASE):
                ip_col = idx
            elif host_col is None and HOST_HEADER_RE.search(col_str):
                host_col = idx
            elif zone_col is None and ZONE_HEADER_RE.search(col_str):
                zone_col = idx
            elif dept_col is None and DEPT_HEADER_RE.search(col_str):
                dept_col = idx
            elif owner_col is None and OWNER_HEADER_RE.search(col_str):
                owner_col = idx
            elif desc_col is None and DESC_HEADER_RE.search(col_str):
                desc_col = idx

        found_ips = []
        asset_map = {}
        meta_map = {}  # ip -> {"zone_tag":.., "department":.., "owner_name":.., "description":..}
        data_rows = rows[1:] if ip_col is not None else rows
        if ip_col is not None:
            for r in data_rows:
                if ip_col < len(r):
                    ip = _looks_like_ip(r[ip_col])
                    if ip:
                        found_ips.append(ip)
                        if host_col is not None and host_col < len(r):
                            hostname = str(r[host_col]).strip()
                            if hostname:
                                asset_map[ip] = hostname
                        meta = {}
                        for field_name, col_idx in (("zone_tag", zone_col), ("department", dept_col),
                                                     ("owner_name", owner_col), ("description", desc_col)):
                            if col_idx is not None and col_idx < len(r):
                                val = str(r[col_idx]).strip()
                                if val:
                                    meta[field_name] = val
                        if meta:
                            meta_map[ip] = meta
        else:
            # 헤더에서 IP 컬럼을 못 찾으면 모든 셀을 훑어 IP처럼 보이는 값을 수집 (수동 검토 필요)
            # - 이 경로는 컬럼 위치를 알 수 없어 asset_map/meta_map을 채우지 않는다.
            for r in rows:
                for cell in r:
                    ip = _looks_like_ip(cell)
                    if ip:
                        found_ips.append(ip)

        # [보안] Command Injection 등으로 이어질 수 있는 값은 여기서도 한 번 더 걸러낸다
        safe_ips = sorted(set(ip for ip in found_ips if OSUtils.is_safe_host(ip)))

        if not safe_ips:
            QMessageBox.warning(self, "Notice", "파일에서 유효한 IP 주소를 찾지 못했습니다.\n'IP' 헤더가 포함된 컬럼이 있는지 확인해주세요.")
            return

        self.ip_input.setText(",".join(safe_ips))
        # [P0] 이번에 가져온 목록을 세션 전체에서 참조할 수 있게 보관 - hostname 폴백과
        # 스캔 종료 후 diff 둘 다 이 값을 쓴다. 새로 가져오기를 하면 이전 목록은 대체된다.
        self.imported_asset_map = {ip: name for ip, name in asset_map.items() if ip in safe_ips}

        # [자산 import 강화] 구역태그/부서/담당자/설명은 스캔을 기다리지 않고 지금
        # 바로 자산을 등록(upsert)해서 반영한다 - 정보자산목록 자체가 이 값들의
        # 권위있는 출처이므로 매번 최신 파일 값으로 덮어쓴다(빈 셀은 건드리지 않음).
        meta_applied = 0
        for ip in safe_ips:
            meta = meta_map.get(ip)
            if not meta:
                continue
            hostname = asset_map.get(ip, "")
            asset_id = self.db.save_asset(ip, hostname=hostname or "Unknown",
                                           hostname_source="매핑" if hostname else "")
            if not asset_id:
                continue
            for field_name, val in meta.items():
                self.db.update_asset_field(asset_id, field_name, val)
            meta_applied += 1

        host_note = f", 호스트명 {len(self.imported_asset_map)}건" if self.imported_asset_map else ""
        meta_note = f", 구역/부서/담당자/설명 {meta_applied}건 반영" if meta_applied else ""
        self.log_message(f"[System] 자산목록 가져오기 완료: {len(safe_ips)}개 IP{host_note}{meta_note} ({os.path.basename(filepath)})")
        QMessageBox.information(self, "가져오기 완료",
            f"{len(safe_ips)}개의 IP를 스캔 대상으로 불러왔습니다.\n(목록에 없는 IP는 스캔 후보에서 제외됩니다)"
            + (f"\n\n호스트명 {len(self.imported_asset_map)}건도 함께 인식했습니다 - DNS로 이름을 못 얻는 자산의 폴백으로 쓰입니다." if self.imported_asset_map else "")
            + (f"\n구역태그/부서/담당자/설명 중 인식된 값을 자산 {meta_applied}건에 바로 반영했습니다." if meta_applied else ""))

    @staticmethod
    def _read_asset_rows_csv(filepath):
        rows = []
        # 정보자산목록은 EUC-KR/CP949로 작성된 경우가 흔해 UTF-8 실패 시 폴백한다
        encodings = ["utf-8-sig", "cp949", "utf-8"]
        last_err = None
        for enc in encodings:
            try:
                with open(filepath, "r", encoding=enc, newline="") as f:
                    rows = [row for row in csv.reader(f) if any(c.strip() for c in row)]
                return rows
            except (UnicodeDecodeError, UnicodeError) as e:
                last_err = e
                continue
        if last_err:
            raise last_err
        return rows

    @staticmethod
    def _read_asset_rows_excel(filepath):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cleaned):
                rows.append(cleaned)
        wb.close()
        return rows

    # [자산 export - 2026-08-06 신규] 자산 목록 CSV/Excel 내보내기.
    EXPORT_HEADERS = ["IP", "Hostname", "Hostname 출처", "OS", "MAC", "Vendor",
                      "구역태그", "부서", "담당자", "설명", "용도", "최종 확인"]

    def export_asset_list(self):
        try:
            assets = self.db.get_all_assets_for_export()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"자산 목록을 불러오지 못했습니다:\n{e}")
            return

        if not assets:
            QMessageBox.warning(self, "Notice", "내보낼 자산이 없습니다.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "자산 목록 내보내기", "asset_list.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not filepath:
            return

        try:
            if filepath.lower().endswith(".csv"):
                # [보안] 엑셀에서 CSV를 열 때 "="로 시작하는 셀은 수식으로 해석돼
                # CSV Injection으로 이어질 수 있다 - 그런 셀 앞에 탭 문자를 붙여 무력화.
                def _csv_safe(val):
                    s = "" if val is None else str(val)
                    return ("\t" + s) if s.startswith(("=", "+", "-", "@")) else s
                with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(self.EXPORT_HEADERS)
                    for row in assets:
                        writer.writerow([_csv_safe(v) for v in row])
            else:
                if not filepath.lower().endswith(".xlsx"):
                    filepath += ".xlsx"
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "자산 목록"
                ws.append(self.EXPORT_HEADERS)
                for row in assets:
                    ws.append(["" if v is None else v for v in row])
                wb.save(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"내보내기 중 오류가 발생했습니다:\n{e}")
            return

        self.log_message(f"[System] 자산 목록 내보내기 완료: {len(assets)}건 -> {filepath}")
        if QMessageBox.question(self, "완료", f"자산 {len(assets)}건을 내보냈습니다.\n파일을 여시겠습니까?",
                                 QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.open_file_platform_safe(filepath)

    def open_db_manager(self):
        from gui.db_manager import DatabaseManagerDialog

        # DB 커넥터 인스턴스를 넘겨줍니다
        dlg = DatabaseManagerDialog(self.db, self)
        dlg.exec()

        # 매니저 창이 닫히면 메인 화면의 통계나 리스트도 갱신하는 것이 좋음
        self.refresh_dashboard() # (만약 이런 기능이 있다면)

    def open_asset_assessment(self):
        from gui.asset_assessment_dialog import AssetAssessmentDialog
        dlg = AssetAssessmentDialog(self.db, self)
        dlg.exec()

    def open_expert_mode(self):
        if not self.license_mgr.can_use_expert_mode():
            QMessageBox.warning(self, "License Restricted",
                "전문가 모드는 Enterprise 등급 전용 기능입니다.\n"
                f"(현재 등급: {self.license_mgr.effective_tier()})")
            return
        from gui.expert_mode_dialog import ExpertModeDialog
        dlg = ExpertModeDialog(self)
        dlg.exec()

    def open_waiver_manager(self):
        # [버그 수정] 예전엔 Waiver에 등급 제한이 아예 없었다 - Expert Mode와 같은
        # "전문가용" 기능으로 통일하면서 Enterprise 전용으로 맞췄다. 버튼 자체가
        # 이제 Enterprise가 아니면 숨겨지지만, 방어적으로 메서드 안에서도 한 번 더 막는다.
        if not self.license_mgr.can_use_waiver():
            QMessageBox.warning(self, "License Restricted",
                "예외처리(Waiver) 관리는 Enterprise 등급 전용 기능입니다.\n"
                f"(현재 등급: {self.license_mgr.effective_tier()})")
            return
        from gui.waiver_dialog import WaiverManagerDialog
        dlg = WaiverManagerDialog(self.db, self)
        dlg.exec()

    def open_manual_audit(self):
        """[수동/육안 점검] 이 경로로 입력된 결과는 TBL_SCAN_RESULT에 자동 스캔과
        동일하게 저장되므로, 닫힌 뒤 대시보드를 갱신한다(open_db_manager와 동일한 이유)."""
        from gui.manual_audit_dialog import ManualAuditDialog
        dlg = ManualAuditDialog(self.db, self)
        dlg.exec()
        self.refresh_dashboard()

    def open_cross_check(self):
        """[교차검증 모드] 스크립트 TXT 결과를 재판정 또는 DB 직접대조로 대조.
        기본(재판정) 모드는 DB에 접근하지 않지만, self.db를 넘겨주면 사용자가 다이얼로그
        안에서 "DB 직접대조" 모드를 선택할 수 있다 - 그 경우에만 실제로 DB를 조회한다.
        어느 모드든 DB에 쓰지는 않으므로 닫힌 뒤 refresh_dashboard()는 불필요하다."""
        from gui.crosscheck_dialog import CrossCheckDialog
        dlg = CrossCheckDialog(self, db=self.db)
        dlg.exec()

    def open_pc_toolkit(self):
        """[PC 진단 도구] 교차검증과 달리 이 경로로 가져온 결과는 TBL_SCAN_RESULT에
        직접 반영되므로, 닫힌 뒤 대시보드를 갱신한다(open_db_manager와 동일한 이유)."""
        from gui.pc_toolkit_dialog import PCToolkitDialog
        dlg = PCToolkitDialog(self, db=self.db)
        dlg.exec()
        self.refresh_dashboard()

    def open_settings(self):
        from gui.settings_dialog import SettingsDialog
        dlg = SettingsDialog(self.db, self)
        if dlg.exec() == QDialog.Accepted:
            from utils.app_settings import load_settings as _load_app_settings
            self._app_settings = _load_app_settings()
            self.log_panel.setVisible(self._app_settings.get("show_log_panel", False))

    def open_help(self):
        from gui.help_dialog import HelpDialog
        dlg = HelpDialog(self)
        dlg.exec()

    def refresh_dashboard(self):
        # [DB Sync] 최신 데이터로 화면 갱신

        # 1. [Phase 3] 대시보드 요약 카드 + 결과 테이블 갱신
        self.refresh_dashboard_metrics()
        self.refresh_findings_table()

        # 2. 테이블 초기화
        self.asset_table.setRowCount(0)
        self.scanned_ip_cache = set()
        
        # 3. 데이터 가져오기 ([버그 수정] hostname을 반환하지 않던 문제 고쳐서 Host 칸이
        #    항상 빈칸이던 걸 해결 - vendor는 여전히 이 쿼리로 조회되지 않아 빈 값)
        assets = self.db.get_all_assets()

        # 화면 깜빡임 방지
        self.asset_table.setSortingEnabled(False)
        self.asset_table.setUpdatesEnabled(False)

        for data in assets:
            try:
                ip, os_type, _description, mac_addr, hostname, hostname_source = data

                # [데이터 세탁] DB에 '-'라고 저장된 것만 보기 좋게 빈칸으로 변경
                # (실제 데이터가 있으면 그대로 유지됨)
                os_type  = "" if str(os_type) == "-"  else os_type
                mac_addr = "" if str(mac_addr) == "-" else mac_addr

                self.add_asset_to_table(ip, hostname, os_type, mac_addr, "", hostname_source)

            except Exception as e:
                print(f"Error refreshing row: {e}")
            
        self.asset_table.setUpdatesEnabled(True)
        self.asset_table.setSortingEnabled(True)
        
        self.log_message(f"[System] Dashboard Refreshed. (Assets: {len(assets)})")
        
    def update_ui_by_license(self):
        # 1. 윈도우 제목 변경
        new_title = self.license_mgr.get_window_title()
        self.setWindowTitle(new_title)

        # [버그 수정] self.title_label은 _build_page_header()가 호출될 때마다 그
        # QLabel로 재바인딩되는 공유 속성이라, initUI() 이후엔 "마지막으로 만들어진
        # 페이지 헤더"(Scan 페이지)만 가리킨다. 여기서 new_title(예: "Z-Vuln Scan v3.0
        # Enterprise")로 덮어쓰면 Scan 페이지 제목이 "Scan"에서 그 문자열로 바뀌는
        # 실측 확인된 부작용이 있었다 - 페이지 제목은 건드리지 않는다.

        # 2. 상단바 칩(담당자·라이선스 등급) 갱신
        self.refresh_topbar_chip()

        # 3. [UI/UX 개선 - 2026-08-06] Waiver/Expert 버튼 표시 여부 - Enterprise가
        # 아니면 아예 숨긴다(클릭 후 차단 메시지 대신). demo_toggle_license()(Ctrl+Shift+L
        # 개발자 미리보기)로 등급을 바꿔도 여기서 다시 호출되므로 즉시 반영된다.
        if hasattr(self, 'btn_waiver'):
            self.btn_waiver.setVisible(self.license_mgr.can_use_waiver())
        if hasattr(self, 'btn_expert'):
            self.btn_expert.setVisible(self.license_mgr.can_use_expert_mode())

        # 4. 엑셀 내보내기는 클릭 시점에 generate_excel()에서 라이선스를 강제하므로
        #    (Reports 메뉴 항목이라 상시 위젯이 없어) 여기서는 별도 처리 없음.

    # [숨겨진 개발자 전용 동작] Ctrl+Shift+L로만 접근 (메뉴/버튼 노출 없음).
    # 등급별 제한이 실제로 어떻게 보이는지 미리보기용 - license.dat는 건드리지 않는다.
    def demo_toggle_license(self):
        new_tier = self.license_mgr.toggle_tier()
        self.update_ui_by_license()
        QMessageBox.information(self, "License Tier Changed (Dev Preview)",
            f"[미리보기] 등급을 {new_tier}로 전환했습니다.\n"
            "(license.dat는 그대로이며, 프로그램을 재시작하면 원래 등급으로 돌아갑니다)")

    def open_license_dialog(self):
        dlg = LicenseDialog(self)
        if dlg.exec():
            if dlg.verified_tier:
                self.license_mgr.current_tier = dlg.verified_tier
                self.license_mgr.expiry_date = dlg.verified_expiry
                self.update_ui_by_license()
                self.log_message(f"[System] License Activated: {dlg.verified_tier} (expires {dlg.verified_expiry})")
        
    def _check_for_updates(self):
        # [버전 업데이트 배포 경로] GitHub Releases(공개 저장소)를 조회한다.
        # 네트워크 실패/릴리즈 없음/이미 최신 버전이면 update_info가 None이라
        # 아무 일도 일어나지 않는다 - 시작 속도에 영향을 주지 않기 위해 짧은
        # timeout(기본 5초)으로 best-effort 조회만 한다.
        update_info = check_for_updates()
        if not update_info:
            return

        box = QMessageBox(self)
        box.setWindowTitle("새 버전 알림")
        box.setText(
            f"새 버전이 있습니다: {update_info.get('latest_version')}\n\n"
            f"{update_info.get('notes', '')}"
        )
        download_btn = box.addButton("다운로드 페이지 열기", QMessageBox.AcceptRole)
        box.addButton("나중에", QMessageBox.RejectRole)
        box.exec()

        if box.clickedButton() == download_btn:
            download_url = update_info.get("download_url", "")
            if download_url:
                QDesktopServices.openUrl(QUrl(download_url))

    def _check_ruleset_update(self):
        # [룰셋 증분 업데이트] 기본 OFF(app_settings.ruleset_auto_update) - 이 함수
        # 자체는 항상 호출하지만, rule_update.check_and_apply_update()가 설정을
        # 직접 확인해서 꺼져 있으면 네트워크 요청 없이 즉시 "DISABLED"로 반환한다
        # (버전 업데이트 확인과 동일하게 시작 흐름에 조용히 얹힌다).
        from utils import rule_update
        status, detail = rule_update.check_and_apply_update()
        if status == "UPDATED":
            self.log_message(f"[RuleUpdate] {detail}")
            QMessageBox.information(self, "룰셋 업데이트 완료", detail)
        elif status == "VERIFY_FAILED":
            # [보안] 서명 검증 실패는 조용히 넘기지 않는다 - 배포 채널이 뭔가
            # 잘못됐다는 신호일 수 있어 사용자가 반드시 인지해야 한다.
            self.log_message(f"[RuleUpdate] 서명 검증 실패: {detail}")
            QMessageBox.warning(self, "룰셋 업데이트 서명 검증 실패", detail)
        # DISABLED/UP_TO_DATE/NO_RELEASE_ASSET/ERROR는 조용히 넘어간다
        # (버전 업데이트 확인과 동일하게 시작 속도/사용자 경험에 영향 안 주도록)

    def sync_to_cloud(self):
        # 아직 기능은 없지만, 사업계획서상 '로드맵' 기능을 시연하는 용도
        QMessageBox.information(self, "Enterprise Feature", 
            "[Cloud Sync]\n\n"
            "자산 데이터를 중앙 관제 대시보드(SaaS)로 전송합니다.\n"
            "(현재 데모 버전에서는 시뮬레이션만 수행됩니다.)")
        
    def load_saved_data(self):
        #[Startup] DB에 저장된 자산 정보를 불러와 테이블에 채웁니다.
        self.log_message("[System] Loading saved assets from database...")
        
        db = DBConnector()
        assets = db.get_all_assets()
        
        if not assets:
            self.log_message("[System] No saved assets found.")
            return

        # 화면 깜빡임 방지 (대량 데이터 로드 시 필수)
        self.asset_table.setSortingEnabled(False)
        self.asset_table.setUpdatesEnabled(False)
        
        count = 0
        for asset in assets:
            try:
                # [버그 수정] hostname을 반환하지 않던 문제 고쳐서 시작 시 Host 칸이
                # 항상 빈칸이던 걸 해결 - vendor는 여전히 이 쿼리로 조회되지 않아 빈 값
                ip, os_type, _description, mac, hostname, hostname_source = asset
                self.add_asset_to_table(ip, hostname, os_type, mac, "", hostname_source)
                count += 1
            except Exception as e:
                print(f"Error loading asset: {e}")
        
        self.asset_table.setUpdatesEnabled(True)
        self.asset_table.setSortingEnabled(True)

        # [Phase 3] 대시보드 요약 카드 갱신
        self.refresh_dashboard_metrics()
        self.refresh_findings_table()
        self.log_message(f"[System] Successfully loaded {count} assets.")