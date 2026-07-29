# --------------------------------------------------------------------------
# Copyright © 2025 Z-VulnScan Team. All Rights Reserved.
#
# This software is proprietary and confidential.
# Unauthorized copying, modification, distribution, or reverse engineering
# of this file, via any medium, is strictly prohibited.
# --------------------------------------------------------------------------
"""
[교차검증 모드] core.crosscheck_engine.run_cross_check()의 결과 dict를 받아
독립 엑셀 워크북으로 출력한다.

excel_report.ExcelGenerator와 스타일 톤은 맞추되(같은 색상/폰트 값을 재정의),
import는 하지 않는 완전 독립 클래스다 - ExcelGenerator는 DB(TBL_SCAN_RESULT)에
강하게 결합되어 있는데, 교차검증 결과는 DB에 전혀 쓰지 않기로 확정했기 때문에
같은 클래스를 재사용하면 오히려 DB 의존을 끌어들이게 된다.
"""
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

from utils.app_settings import get_report_output_dir  # noqa: E402
from core.crosscheck_engine import MISSING_CODE_LIMITATION_NOTE  # noqa: E402


class CrossCheckReportGenerator:
    COLOR_HEADER_BG = 'E7E6E6'
    COLOR_MATCH = 'E2EFDA'      # 일치 (초록)
    COLOR_MISMATCH = 'FFCCCC'   # 불일치 (빨강)
    COLOR_MISSING = 'FFF2CC'    # 누락/판별불가 (노랑)
    COLOR_ERROR = 'D9D9D9'      # 파싱 실패 (회색)

    FONT_TITLE = Font(name='맑은 고딕', size=16, bold=True)
    FONT_HEADER = Font(name='맑은 고딕', size=10, bold=True)
    FONT_NORMAL = Font(name='맑은 고딕', size=9)
    FONT_BOLD = Font(name='맑은 고딕', size=9, bold=True)

    BORDER_THIN = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    _CLASS_COLOR = {
        "MATCH": COLOR_MATCH,
        "MISMATCH": COLOR_MISMATCH,
        "MISSING_IN_RULESET": COLOR_MISSING,
        "AMBIGUOUS_RULESET": COLOR_MISSING,
        "PARSE_ERROR": COLOR_ERROR,
        "NO_KISA_CODE": COLOR_ERROR,
        "NO_CONSULTANT_RESULT": COLOR_ERROR,
        "CONSULTANT_UNCERTAIN": COLOR_MISSING,
        "MANUAL_VERIFICATION_NEEDED": COLOR_MISSING,
        "DB_NO_ASSET": COLOR_ERROR,
        "DB_NO_SCAN_RESULT": COLOR_ERROR,
    }
    _CLASS_LABEL = {
        "MATCH": "일치",
        "MISMATCH": "불일치",
        "MISSING_IN_RULESET": "룰셋에 코드 없음",
        "AMBIGUOUS_RULESET": "룰셋 판별 불가",
        "PARSE_ERROR": "파싱 실패",
        "NO_KISA_CODE": "KISA 코드 없음",
        "NO_CONSULTANT_RESULT": "스크립트 결과 없음",
        "CONSULTANT_UNCERTAIN": "스크립트도 확인 필요",
        "MANUAL_VERIFICATION_NEEDED": "수동 검증 필요(세부기준형)",
        "DB_NO_ASSET": "Z-VulnScan 스캔 자산 없음",
        "DB_NO_SCAN_RESULT": "Z-VulnScan 스캔 이력 없음",
    }

    def __init__(self, cross_check_result: dict):
        """DB 의존 없음 - run_cross_check()의 반환 dict만 입력받는다."""
        self.result = cross_check_result
        self.output_dir = get_report_output_dir()
        if not os.path.exists(self.output_dir):
            try:
                os.makedirs(self.output_dir, exist_ok=True)
            except OSError:
                pass
        self.header_fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type='solid')

    def _write_header(self, ws, headers, row=1):
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = self.FONT_HEADER
            cell.fill = self.header_fill
            cell.border = self.BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _autofit(self, ws, widths):
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _write_summary_sheet(self, wb):
        ws = wb.active
        ws.title = "Summary"
        s = self.result.get("summary", {})

        ws["A1"] = "교차검증 결과 요약"
        ws["A1"].font = self.FONT_TITLE
        ws["A2"] = f"생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = self.FONT_NORMAL

        rows = [
            ("총 점검 항목 수", s.get("total", 0)),
            ("일치 (스크립트 판정 = 재판정)", s.get("match", 0)),
            ("불일치", s.get("mismatch", 0)),
            ("룰셋에 코드 없음", s.get("missing_in_ruleset", 0)),
            ("룰셋 판별 불가 (D-xx MySQL/PostgreSQL)", s.get("ambiguous", 0)),
            ("파싱 실패 (수동 확인 필요)", s.get("parse_error", 0)),
            ("근사치 재판정 항목 (세부기준형 룰)", s.get("approx_count", 0)),
            ("호스트별 누락 코드 수", s.get("missing_codes", 0)),
            ("KISA 코드 없음 (스크립트 측 매핑 불가 항목)", s.get("no_kisa_code", 0)),
            ("스크립트 결과 없음 (수동 확인 항목 등)", s.get("no_consultant_result", 0)),
            ("스크립트도 확인 필요 (\"확인\" 판정)", s.get("consultant_uncertain", 0)),
            ("수동 검증 필요 (세부기준형 룰, 자동 대조 신뢰 불가)", s.get("manual_verification_needed", 0)),
            ("[DB 직접대조 모드] Z-VulnScan 스캔 자산 없음", s.get("db_no_asset", 0)),
            ("[DB 직접대조 모드] Z-VulnScan 스캔 이력 없음", s.get("db_no_scan_result", 0)),
        ]
        r = 4
        for label, value in rows:
            ws.cell(row=r, column=1, value=label).font = self.FONT_BOLD
            ws.cell(row=r, column=2, value=value).font = self.FONT_NORMAL
            r += 1

        total = s.get("total", 0)
        match = s.get("match", 0)
        match_rate = f"{(match / total * 100):.1f}%" if total else "N/A"
        r += 1
        ws.cell(row=r, column=1,
                value=f"총 {total}개 항목 중 {match}개 일치 ({match_rate}) - 교차검증 완료").font = self.FONT_BOLD
        r += 2
        ws.cell(row=r, column=1, value="[한계 고지]").font = self.FONT_BOLD
        r += 1
        ws.cell(row=r, column=1, value=MISSING_CODE_LIMITATION_NOTE).font = self.FONT_NORMAL
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1
        ws.cell(row=r, column=1,
                value="'N/A' 판정은 스크립트 원본에서도 해당없음/수동확인/접속실패가 모두 동일 문자열로 "
                      "기록되므로(Z-VulnScan 자체 출력 포맷의 알려진 제약), 세부 원인까지는 구분되지 않습니다.").font = self.FONT_NORMAL
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

        self._autofit(ws, [50, 20])

    def _write_diff_rows(self, ws, entries, start_row=2):
        r = start_row
        for e in entries:
            values = [
                e.get("host"), e.get("ip"), e.get("code"), e.get("name"), e.get("category"),
                e.get("consultant_result"), e.get("rejudged_result"),
                self._CLASS_LABEL.get(e.get("classification"), e.get("classification")),
                e.get("approx_note") or "", e.get("detail_note") or "",
            ]
            color = self._CLASS_COLOR.get(e.get("classification"))
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid') if color else None
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_THIN
                if fill:
                    cell.fill = fill
            r += 1
        return r

    def _write_mismatch_sheet(self, wb):
        ws = wb.create_sheet("Mismatch")
        headers = ["Host", "IP", "Code", "Name", "Category", "스크립트 판정", "재판정", "구분", "근사치", "비고"]
        self._write_header(ws, headers)
        mismatches = [e for e in self.result.get("diff_entries", []) if e["classification"] == "MISMATCH"]
        self._write_diff_rows(ws, mismatches)
        self._autofit(ws, [18, 15, 10, 30, 16, 12, 12, 16, 30, 40])

    def _write_missing_sheet(self, wb):
        ws = wb.create_sheet("Missing")
        headers = ["Host", "IP", "Ruleset", "Code", "Name"]
        self._write_header(ws, headers)
        r = 2
        for m in self.result.get("missing_entries", []):
            values = [m.get("host"), m.get("ip"), m.get("ruleset_file"), m.get("code"), m.get("name")]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_THIN
                fill = PatternFill(start_color=self.COLOR_MISSING, end_color=self.COLOR_MISSING, fill_type='solid')
                cell.fill = fill
            r += 1
        ws.cell(row=r + 1, column=1, value=MISSING_CODE_LIMITATION_NOTE).font = self.FONT_NORMAL
        self._autofit(ws, [18, 15, 20, 10, 30])

    def _write_all_items_sheet(self, wb):
        ws = wb.create_sheet("All Items")
        headers = ["Host", "IP", "Code", "Name", "Category", "스크립트 판정", "재판정", "구분", "근사치", "비고"]
        self._write_header(ws, headers)
        self._write_diff_rows(ws, self.result.get("diff_entries", []))
        self._autofit(ws, [18, 15, 10, 30, 16, 12, 12, 16, 30, 40])

    def generate(self, output_path=None):
        wb = Workbook()
        self._write_summary_sheet(wb)
        self._write_mismatch_sheet(wb)
        self._write_missing_sheet(wb)
        self._write_all_items_sheet(wb)

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"[CrossCheck]_{timestamp}.xlsx")

        wb.save(output_path)
        return output_path
