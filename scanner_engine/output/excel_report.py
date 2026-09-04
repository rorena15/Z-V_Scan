# --------------------------------------------------------------------------
# Copyright © 2025 Z-VulnScan Team. All Rights Reserved.
#
# This software is proprietary and confidential.
# Unauthorized copying, modification, distribution, or reverse engineering
# of this file, via any medium, is strictly prohibited.
# --------------------------------------------------------------------------
"""
[리포트 재설계] 카테고리(UNIX/WINDOWS/PC/웹서비스/DBMS)별로 요약통계/보안수준통계
시트 + 호스트 1대당 상세 시트 1개 구조로 구성한다. 카테고리별로 파일을 따로 만들지
않고, Z-VulnScan은 파일 하나(통합 워크북) 안에서 시트로만 구분한다(사용자 결정).
관리적/물리적/보안장비/네트워크장비 영역은 Z-VulnScan이 자동 점검하는 영역이
아니라 이번 재설계 범위에서 제외했다.

점수 산식(항목별 중요도 기반 가중 감점 방식):
  중요도 값: 상=10, 중=8, 하=6
  결과 값(항목 하나의 감점): 양호=0, 부분만족=중요도값/2, 취약=중요도값, 해당없음은 제외
  보안수준 = (전체점수-취약점수)/전체점수   (전체점수=해당없음 제외 항목의 중요도 값 합계)
매크로가 없는 순수 openpyxl 산출물이라 수식이 아니라 계산된 값을 셀에 직접 쓴다.

서식(색상)은 일반적인 KISA 점검 결과보고서 스타일의 회색조 대신, Z-VulnScan
데스크톱 앱과 동일한 브랜드 팔레트(gui/dashboard_widgets.py의 LIGHT_COLORS)를
쓴다 - 리포트만 봐도 Z-VulnScan 산출물임을 알아볼 수 있게 하기 위함(사용자 지시:
"서식을 완전히 똑같이 해야할 필요는 없지만 Z-Vuln Scan만의 색으로").
"""
import os
import sys
import sqlite3
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# 상위 모듈 참조
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from utils.db_connector import DBConnector
from utils.logger import AppLogger
from utils.app_settings import get_report_output_dir
from utils import rule_crypto
import core.vuln_matcher as Vuln
from output.text_report import RULE_FILES, STATUS_TO_RESULT
# [버그 수정] IMPORTANCE_WEIGHT가 이 파일/db_connector.py 두 곳에 따로 복사돼
# 있어 하나만 안 고치면 조용히 어긋났다 - 공용으로 통합(RULE_FILES는 이미
# text_report.py를 통해 간접 공유되고 있었으므로 그대로 둠. 새 파일을 만들지
# 않고 rule_crypto.py에 합침 - PyArmor 트라이얼 파일 수 제약).
from utils.rule_crypto import IMPORTANCE_WEIGHT

# 코드 접두어 -> 표시 카테고리명. 순서가 그대로 시트 등장 순서가 된다.
# 관리적/물리적/보안장비/네트워크장비는 자동 점검 대상이 아니라 제외.
CATEGORY_PREFIXES = [
    ("UNIX", ("U-",)),
    ("WINDOWS", ("W-",)),
    ("PC", ("PC-",)),
    ("웹서비스", ("WEB-",)),
    ("DBMS", ("MYSQL-", "POSTGRESQL-", "MSSQL-", "ORACLE-")),
]


def _categorize_code(code):
    """kisa_code(없으면 vuln_code) 접두어로 카테고리를 판별한다.
    매칭 안 되면(Discovery 정보 TCP-/UDP-/INFO- 등) None."""
    if not code:
        return None
    for cat_name, prefixes in CATEGORY_PREFIXES:
        if any(code.startswith(p) for p in prefixes):
            return cat_name
    return None


class ExcelGenerator:
    # --- Z-VulnScan 브랜드 팔레트 (gui/dashboard_widgets.py LIGHT_COLORS와 동일 값) ---
    COLOR_HEADER_BG = 'E3ECFE'   # accent_bg - 헤더/구분선 배경
    COLOR_ACCENT = '1F5FE0'      # accent - 제목/강조 텍스트
    COLOR_BORDER = 'D8D8D2'
    COLOR_TEXT = '1A1A1A'

    COLOR_CRITICAL = 'F8C6C6'  # danger_bg - 취약
    COLOR_GOOD = 'C9E9D1'      # success_bg - 양호
    COLOR_PARTIAL = 'F3D27E'   # warning_bg - 부분만족/검토필요
    COLOR_NA = 'E7E7E4'        # muted_bg - 해당없음/점검불가
    COLOR_WAIVER = 'D8D8D2'    # border색 재사용 - 예외

    FONT_TITLE = Font(name='맑은 고딕', size=20, bold=True, color=COLOR_ACCENT)
    FONT_SECTION = Font(name='맑은 고딕', size=12, bold=True, color=COLOR_ACCENT)
    FONT_HEADER = Font(name='맑은 고딕', size=10, bold=True, color=COLOR_TEXT)
    FONT_NORMAL = Font(name='맑은 고딕', size=9)
    FONT_BOLD = Font(name='맑은 고딕', size=9, bold=True)

    BORDER_ALL = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )

    STATUS_FILL = {
        "취약": COLOR_CRITICAL, "양호": COLOR_GOOD, "부분만족": COLOR_PARTIAL,
        "해당없음": COLOR_NA, "예외": COLOR_WAIVER, "검토필요": COLOR_PARTIAL,
        "점검불가": COLOR_NA,
    }

    def __init__(self, evidence_level="full", remediation_level="full",
                 report_title=None, company_name=None, custom_filename=None,
                 asset_ids=None, codes=None):
        self.db = DBConnector()
        # [라이선스 등급별 리포트 차등] evidence_level: "full"|"vuln_only"
        # remediation_level: "full"|"partial"|"none" - 기존과 동일하게 유지
        self.evidence_level = evidence_level
        self.remediation_level = remediation_level
        # [리포트 탭 - 자산별/항목별 선택, 2026-09] None(기본값)이면 지금까지처럼
        # 전체 자산/전체 항목을 담는다 - 하위호환. _fetch_scan_result()/
        # _fetch_asset_assessment_full()에서만 걸러서 카테고리/호스트 시트 등
        # 이 둘에서 파생되는 모든 내용에 자동으로 반영되게 한다. 회차 비교(diff)/
        # 점검대상 커버리지 시트는 이번 스캔 범위가 아니라 누적 이력 성격이라
        # 의도적으로 필터를 적용하지 않는다.
        self.asset_ids = list(asset_ids) if asset_ids else None
        self.codes = list(codes) if codes else None
        # [리포트 탭 - 커스터마이징] 표지 제목/브랜딩 문구/저장 파일명 - 비워두면
        # 지금까지와 동일한 기본값(PDFGenerator와 동일한 정책).
        self.report_title = (report_title or "").strip() or "주요정보통신기반시설 기술적 취약점 분석 평가 결과"
        self.company_name = (company_name or "").strip()
        self.custom_filename = re.sub(r'[\\/*?:"<>|\[\]]', '_', (custom_filename or "").strip())[:80]
        self.output_dir = get_report_output_dir()
        if not os.path.exists(self.output_dir):
            try:
                os.makedirs(self.output_dir, exist_ok=True)
            except OSError:
                pass

        self.HEADER_FILL = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type='solid')
        self.rule_lookup = self._load_rule_lookup()
        self._used_sheet_names = set()

    # ------------------------------------------------------------------
    # 공용 유틸 (기존 excel_report.py에서 그대로 유지)
    # ------------------------------------------------------------------
    def clean_text(self, text):
        # [버그 수정] "if not text"는 파이썬에서 0/0.0도 falsy라 실제 값이 0인
        # 취약건수·CIA 점수 등이 "누락"과 똑같이 "-"로 표시되고 있었다. None/빈
        # 문자열만 "없음"으로 취급한다.
        if text is None or text == "":
            return "-"
        text = str(text)
        text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
        # [버그 수정] Excel 수식 인젝션 방지 - raw_output/detected_value는 스캔
        # 대상이 실제로 응답한 값이라 신뢰할 수 없는 입력이다. '='/'+'/'-'/'@'로
        # 시작하면 openpyxl이 셀을 수식으로 저장해, 리포트를 여는 사람 PC에서
        # 그대로 실행될 수 있다(OWASP CSV Injection과 동일한 유형). 선행 문자를
        # 작은따옴표로 escape해 값은 보존하되 수식으로 해석되지 않게 한다.
        if text and text[0] in ('=', '+', '-', '@'):
            text = "'" + text
        return text

    def _style_range(self, ws, cell_range, border=None, fill=None, font=None, alignment=None):
        selection = ws[cell_range]
        if not isinstance(selection, (tuple, list)):
            selection = ((selection,),)
        for row in selection:
            for cell in row:
                if border: cell.border = border
                if fill: cell.fill = fill
                if font: cell.font = font
                if alignment: cell.alignment = alignment

    def _add_security_level_chart(self, ws, title, anchor, header_row, first_row, last_row, value_col, cat_col=1):
        """[2026-09] 표지/카테고리 시트/호스트 시트 3곳이 전부 "분류(또는 카테고리)별
        보안수준(%) 막대 차트"를 쓰므로 공용 헬퍼 하나로 합쳤다.
        [범위 버그 수정] y_axis.scaling.min/max를 안 정해두면 Excel이 데이터
        최소~최대로 축을 자동 축소해서(예: 55~70%대 값이면 그 구간만 보여줌) 막대
        높이 차이가 실제보다 과장돼 보인다 - 보안수준은 의미상 0~100% 범위이므로
        축 범위를 항상 명시적으로 고정한다."""
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.y_axis.title = "보안수준"
        chart.y_axis.numFmt = '0%'
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.height = 7
        chart.width = 16
        chart.style = 10
        chart.legend = None
        data_ref = Reference(ws, min_col=value_col, min_row=header_row, max_row=last_row)
        cats_ref = Reference(ws, min_col=cat_col, min_row=first_row, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, anchor)

    def _unique_sheet_name(self, name):
        """엑셀 시트명 31자 제한 + 중복 방지(호스트명이 같거나 잘려서 겹치는 경우)."""
        name = re.sub(r'[\\/*?:\[\]]', '_', name)[:31]
        base, n = name, 2
        while name in self._used_sheet_names:
            suffix = f"_{n}"
            name = base[:31 - len(suffix)] + suffix
            n += 1
        self._used_sheet_names.add(name)
        return name

    def _final_status(self, status, waived):
        """DB status(+예외처리 여부) -> 화면 표시용 4단계+2 상태 문자열.
        기존 excel_report.py의 매핑을 그대로 유지."""
        if waived == 1: return "예외"
        if status == "VULNERABLE": return "취약"
        if status == "PARTIAL": return "부분만족"
        if status == "NA": return "해당없음"
        if status == "MANUAL": return "검토필요"
        if status == "ERROR": return "점검불가"
        return "양호"

    def _load_rule_lookup(self):
        """{엔진접두어}{code} -> rule dict. output/text_report.py._load_rule_lookup()와
        동일 규칙(RULE_FILES) - database_inspector.py가 DB에 저장하는 vuln_code
        (예: MYSQL-D-01)와 키가 일치해야 한다. 새 파싱 로직을 만들지 않고 그 규칙만
        재사용한다."""
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            # [버그 수정] output/text_report.py의 TextReportGenerator와 동일하게 3단계
            # dirname이 필요하다(__file__ -> output/ -> scanner_engine/ -> 프로젝트 루트).
            # 2단계로 계산하면 scanner_engine/rules를 찾게 되는데 실제 rules/는 그 상위
            # 프로젝트 루트에 있어서 룰 조회가 항상 실패(빈 lookup)하던 버그가 있었다.
            base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        rules_dir = os.path.join(base_path, 'rules')
        if hasattr(sys, '_MEIPASS') and not os.path.isdir(rules_dir):
            rules_dir = os.path.join(sys._MEIPASS, 'rules')

        lookup = {}
        for filename, prefix in RULE_FILES.items():
            path = os.path.join(rules_dir, filename)
            if not os.path.exists(path) and not os.path.exists(rule_crypto.get_enc_path(path)):
                continue
            try:
                data = rule_crypto.load_ruleset(path)
                for rule in data:
                    lookup[f"{prefix}{rule['code']}"] = rule
            except Exception as e:
                AppLogger.log_error(f"[ExcelReport] Failed to load {filename}", e)
        return lookup

    def _rule_for(self, code):
        return self.rule_lookup.get(code)

    # ------------------------------------------------------------------
    # 메인 진입점
    # ------------------------------------------------------------------
    def generate(self):
        try:
            wb = Workbook()
            Vuln.VulnMatcher.sync_kisa_code_on_db(self.db.db_path)

            scan_data = self._fetch_scan_result()
            asset_data = self._fetch_asset_assessment_full()

            if not asset_data:
                raise Exception("진단된 자산 데이터가 없습니다.")

            by_category = {cat: [] for cat, _ in CATEGORY_PREFIXES}
            for row in scan_data:
                cat = _categorize_code(row["code"])
                if cat:
                    by_category[cat].append(row)

            self._create_cover_sheet(wb, asset_data, by_category)
            self._create_scope_sheet(wb, asset_data)

            # [2026-09 - "시트가 너무 많다" 피드백 반영] 예전엔 카테고리마다
            # 요약통계/보안수준통계/상세 3개 시트씩 만들었다(5개 카테고리 기준
            # 15개) - 요약통계+보안수준통계를 "{cat}.현황" 하나로 합쳐서 카테고리당
            # 2개 시트(현황+상세)로 줄였다. 요약류 시트끼리, 상세 시트끼리 각각
            # 모아서 배치하는 기존 순서(감사보고서 관례 - 요약을 앞으로)는 유지.
            active_categories = [(cat, by_category[cat]) for cat, _ in CATEGORY_PREFIXES if by_category[cat]]

            for cat_name, rows in active_categories:
                self._create_category_sheet(wb, cat_name, rows)
            for cat_name, rows in active_categories:
                self._create_host_sheets(wb, cat_name, rows)

            round_comparison = self.db.get_round_comparison()
            if round_comparison:
                self._create_diff_sheet(wb, round_comparison)

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"{self.custom_filename}_{timestamp}" if self.custom_filename else f"주요정보통신기반시설_취약점진단결과_{timestamp}"
            filename = f"{base_name}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            return filepath

        except Exception as e:
            AppLogger.log_error("Excel Generation Error", e)
            raise e

    def generate_asset_ledger(self):
        """[자산관리대장, 2026-09 신규] 취약점 진단 결과 워크북(generate())과는
        별개의 독립 문서 - 자산 탭에 이미 있던 "내보내기"(get_all_assets_for_export,
        CSV/Excel 원시 덤프, C/I/A·등급 미포함)와 달리 이 앱의 리포트 서식/색상을
        입힌 정식 자산 등록대장이다. self.asset_ids가 지정돼 있으면(리포트 탭
        범위 선택) 그 자산만 담는다 - self.codes(항목별 선택)는 자산 자체를
        추리는 게 아니라 findings에만 의미가 있어 여기엔 적용하지 않는다."""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            rows = self.db.get_asset_ledger_rows(asset_ids=self.asset_ids)
            self._create_asset_ledger_sheet(wb, rows)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"{self.custom_filename}_자산관리대장_{timestamp}" if self.custom_filename else f"자산관리대장_{timestamp}"
            filename = f"{base_name}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            return filepath
        except Exception as e:
            AppLogger.log_error("Asset Ledger Generation Error", e)
            raise e

    def _create_asset_ledger_sheet(self, wb, rows):
        ws = wb.create_sheet("자산관리대장")
        headers = [
            ("A", "No", 5), ("B", "IP 주소", 15), ("C", "Hostname", 20), ("D", "OS", 18),
            ("E", "MAC", 16), ("F", "Vendor", 16), ("G", "구역태그", 12),
            ("H", "용도", 18), ("I", "부서", 14), ("J", "담당자", 12),
            ("K", "C", 6), ("L", "I", 6), ("M", "A", 6), ("N", "등급", 8),
            ("O", "설명", 24), ("P", "최종 확인", 18),
        ]
        for col, title, width in headers:
            cell = ws[f'{col}1']
            cell.value = title
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        for idx, (asset_id, ip, hostname, os_type, mac, vendor, zone_tag,
                  purpose, dept, owner, c, i, a, description, last_seen) in enumerate(rows, 1):
            grade = DBConnector.compute_asset_grade(c, i, a)
            vals = [
                idx, ip, hostname or "-", os_type or "-", mac or "-", vendor or "-", zone_tag or "-",
                purpose or "-", dept or "-", owner or "-",
                c if c is not None else "-", i if i is not None else "-", a if a is not None else "-", grade,
                description or "-", last_seen or "-",
            ]
            r_idx = idx + 1
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = self.clean_text(val)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # ------------------------------------------------------------------
    # DB 조회
    # ------------------------------------------------------------------
    def _fetch_scan_result(self):
        """카테고리/호스트별 시트 구성에 필요한 필드를 딕셔너리 레코드 리스트로 반환한다."""
        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()
        try:
            filter_sql = ""
            params = []
            if self.asset_ids:
                placeholders = ",".join("?" * len(self.asset_ids))
                filter_sql += f" AND R.asset_id IN ({placeholders})"
                params += list(self.asset_ids)
            if self.codes:
                placeholders = ",".join("?" * len(self.codes))
                filter_sql += (
                    " AND (CASE WHEN R.kisa_code IS NOT NULL AND R.kisa_code != '' "
                    f"THEN R.kisa_code ELSE R.vuln_code END) IN ({placeholders})"
                )
                params += list(self.codes)

            query = f"""
                SELECT
                    A.asset_id, A.ip_addr, A.hostname, A.os_type,
                    A.asset_purpose, A.department, A.owner_name,
                    A.conf_score, A.integ_score, A.avail_score,
                    R.kisa_code, R.vuln_code, R.vuln_name,
                    R.risk_level, R.status, R.detected_value,
                    R.raw_output, R.remediation, R.waiver_status, R.waiver_reason
                FROM TBL_SCAN_RESULT R
                JOIN TBL_ASSETS A ON R.asset_id = A.asset_id
                WHERE R.vuln_code NOT LIKE 'SYS-%'
                AND {DBConnector.latest_round_condition('R')}
                {filter_sql}
                ORDER BY A.ip_addr ASC, R.kisa_code ASC
            """
            cursor.execute(query, params)
            cols = [d[0] for d in cursor.description]
            rows = []
            for raw in cursor.fetchall():
                rec = dict(zip(cols, raw))
                rec["code"] = rec["kisa_code"] or rec["vuln_code"]
                rows.append(rec)
            return rows
        except Exception as e:
            AppLogger.log_error("Fetch Scan Result Error", e)
            return []
        finally:
            conn.close()

    def _fetch_asset_assessment_full(self):
        """점검대상 시트용: 자산평가 필드를 포함한 자산 전체 목록(스캔 결과 유무와 무관).
        [리포트 탭 - 자산별 선택] asset_ids가 있으면 그 자산만 - codes(항목별 선택)는
        자산 자체를 빼지 않으므로 여기엔 적용하지 않는다(findings만 걸러짐)."""
        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()
        try:
            sql = """
                SELECT asset_id, ip_addr, hostname, os_type,
                       asset_purpose, department, owner_name,
                       conf_score, integ_score, avail_score
                FROM TBL_ASSETS
            """
            params = []
            if self.asset_ids:
                placeholders = ",".join("?" * len(self.asset_ids))
                sql += f" WHERE asset_id IN ({placeholders})"
                params = list(self.asset_ids)
            sql += " ORDER BY ip_addr ASC"
            cursor.execute(sql, params)
            return cursor.fetchall()
        except Exception as e:
            AppLogger.log_error("Fetch Asset Assessment Error", e)
            return []
        finally:
            conn.close()

    # ------------------------------------------------------------------
    # 점수 계산 (중요도 가중 감점 방식)
    # ------------------------------------------------------------------
    def _score_group(self, items):
        """items: [(importance, final_status), ...] 하나의 분류(또는 카테고리) 묶음.
        반환: dict(전체, 점검, 양호, 취약, 부분만족, 해당없음, 전체점수, 취약점수, 보안수준)"""
        counts = {"전체": len(items), "양호": 0, "취약": 0, "부분만족": 0, "해당없음": 0}
        full_score = 0
        vuln_score = 0
        for importance, status in items:
            weight = IMPORTANCE_WEIGHT.get(importance, 8)
            if status == "취약":
                counts["취약"] += 1
                full_score += weight
                vuln_score += weight
            elif status == "부분만족":
                counts["부분만족"] += 1
                full_score += weight
                vuln_score += weight / 2
            elif status == "양호":
                counts["양호"] += 1
                full_score += weight
            else:
                # 해당없음/검토필요/점검불가/예외 - 점수 계산에서 제외
                counts["해당없음"] += 1
        counts["점검"] = counts["전체"] - counts["해당없음"]
        counts["전체점수"] = full_score
        counts["취약점수"] = vuln_score
        counts["보안수준"] = None if full_score == 0 else (full_score - vuln_score) / full_score
        return counts

    # ------------------------------------------------------------------
    # 시트: 1.종합요약
    # ------------------------------------------------------------------
    def _create_cover_sheet(self, wb, asset_data, by_category):
        ws = wb.active
        ws.title = "1.종합요약"

        ws.merge_cells('B2:I3')
        ws['B2'] = self.report_title
        self._style_range(ws, 'B2:I3', border=self.BORDER_ALL, font=self.FONT_TITLE,
                           alignment=Alignment(horizontal='center', vertical='center'))

        # [리포트 탭 - 커스터마이징] 회사명/브랜딩 문구 - 지정 안 하면 표시 안 함.
        # 결재란(K~N열, 2~5행)과 겹치지 않는 4행에 넣는다(제목박스는 2~3행만 씀).
        if self.company_name:
            ws.merge_cells('B4:I4')
            ws['B4'] = self.company_name
            self._style_range(ws, 'B4:I4', font=self.FONT_SECTION,
                               alignment=Alignment(horizontal='center', vertical='center'))

        start_col = 11
        col_letter = get_column_letter(start_col)
        range_str = f'{col_letter}2:{col_letter}5'
        ws.merge_cells(range_str)
        ws[f'{col_letter}2'] = "결\n\n재"
        self._style_range(ws, range_str, border=self.BORDER_ALL,
                           fill=self.HEADER_FILL,
                           alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))

        roles = ["담 당", "팀 장", "부서장"]
        for i, role in enumerate(roles):
            c = get_column_letter(start_col + 1 + i)
            ws[f'{c}2'] = role
            self._style_range(ws, f'{c}2', border=self.BORDER_ALL, fill=self.HEADER_FILL,
                               alignment=Alignment(horizontal='center', vertical='center'))
            range_sign = f'{c}3:{c}5'
            ws.merge_cells(range_sign)
            self._style_range(ws, range_sign, border=self.BORDER_ALL)

        row = 7
        ws[f'B{row}'] = "■ 진단 개요"
        ws[f'B{row}'].font = self.FONT_SECTION
        row += 1

        overview_data = [
            ["진단 기간", datetime.now().strftime("%Y년 %m월 %d일")],
            ["진단 대상", f"총 {len(asset_data)}대"],
            ["진단 도구", "Z-VulnScan Pro v3.0 (KISA Guide Compliance)"],
            ["수행자", "자체 보안 점검"]
        ]
        for title, content in overview_data:
            ws[f'B{row}'] = title
            self._style_range(ws, f'B{row}', border=self.BORDER_ALL, fill=self.HEADER_FILL,
                               alignment=Alignment(horizontal='center', vertical='center'))
            range_content = f'C{row}:E{row}'
            ws.merge_cells(range_content)
            ws[f'C{row}'] = content
            self._style_range(ws, range_content, border=self.BORDER_ALL,
                               alignment=Alignment(horizontal='left', vertical='center', indent=1))
            row += 1

        # 카테고리별 보안수준 요약 (기존의 하드코딩 감점식 대신 위 _score_group 산식 사용)
        row += 2
        ws[f'B{row}'] = "■ 카테고리별 보안수준"
        ws[f'B{row}'].font = self.FONT_SECTION
        row += 1

        headers = ["카테고리", "점검항목", "양호", "취약", "부분만족", "해당없음", "보안수준"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2 + i)
            cell.value = h
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center')
        header_row = row
        row += 1

        for cat_name, _ in CATEGORY_PREFIXES:
            cat_rows = by_category.get(cat_name, [])
            if not cat_rows:
                continue
            items = [(self._rule_for(r["code"]).get("importance", "중") if self._rule_for(r["code"]) else "중",
                      self._final_status(r["status"], r["waiver_status"])) for r in cat_rows]
            stats = self._score_group(items)
            values = [
                cat_name, stats["점검"], stats["양호"], stats["취약"], stats["부분만족"], stats["해당없음"],
                "N/A" if stats["보안수준"] is None else f"{stats['보안수준']*100:.0f}%",
            ]
            for i, v in enumerate(values):
                cell = ws.cell(row=row, column=2 + i)
                cell.value = v
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(horizontal='center')
                cell.font = self.FONT_BOLD if i == 0 else self.FONT_NORMAL
            row += 1

        for i in range(len(headers)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 16
        ws.row_dimensions[header_row].height = 20

        # [그래프 추가 - 2026-09 "너무 단조롭다" 피드백] 표지에서 바로 보이도록
        # 카테고리별 보안수준(%) 막대 차트를 표 옆에 붙인다. "보안수준" 열은
        # values 리스트에서 문자열("N/A" 또는 "72%")로 만들어 셀에 쓰지만,
        # 여기서는 표 대신 별도로 숫자(0~1)를 써서 차트가 값을 인식하게 한다 -
        # 표시용 텍스트와 차트용 숫자를 같은 셀에 담을 수 없어서 표 렌더링
        # 로직은 그대로 두고 차트만 별도 보조 열(숨김)에 숫자를 채워 넣는다.
        cat_data_first_row = header_row + 1
        cat_data_last_row = row - 1
        if cat_data_last_row >= cat_data_first_row:
            helper_col = 2 + len(headers)  # 표 바로 오른쪽 열을 숨겨서 숫자 보조열로 사용
            helper_letter = get_column_letter(helper_col)
            ws.cell(row=header_row, column=helper_col, value="보안수준(차트용)")
            for r in range(cat_data_first_row, cat_data_last_row + 1):
                pct_text = ws.cell(row=r, column=2 + len(headers) - 1).value  # "보안수준" 열("72%" 또는 "N/A")
                numeric = None
                if isinstance(pct_text, str) and pct_text.endswith('%'):
                    try:
                        numeric = float(pct_text.rstrip('%')) / 100
                    except ValueError:
                        numeric = None
                ws.cell(row=r, column=helper_col, value=numeric)
            ws.column_dimensions[helper_letter].hidden = True

            # [2026-09 - "한눈에 보기 편하게" 위치 조정] 표 "아래"가 아니라
            # 표 "옆"(helper 열 I 다음, 한 칸 띄운 K열)에 붙인다 - 카테고리/호스트
            # 시트 차트와 같은 배치 규칙이라 문서 전체가 일관되고, 표와 차트를
            # 스크롤 없이 한 화면에서 같이 볼 수 있다. 아래에 두던 이전 방식은
            # 차트 높이만큼(17행) 다음 섹션을 억지로 밀어내야 해서 표와 차트
            # 사이에 큰 빈 공간이 생겼는데, 옆에 붙이면 그 여백 자체가 없어진다.
            self._add_security_level_chart(
                ws, "카테고리별 보안수준", f"K{header_row}",
                header_row=header_row, first_row=cat_data_first_row, last_row=cat_data_last_row,
                value_col=helper_col, cat_col=2,
            )

        # [커버리지 추적] 발견됐지만 KISA 판정이 하나도 없는 자산을 감사 완결성
        # 근거자료로 명시한다 - "몇 %를 봤는가"가 아니라 "빠진 게 있는가"를 정직하게
        # 남기는 게 목적이라, 0건이어도 "완전 일치" 문구를 남겨 침묵하지 않는다.
        row += 2
        ws[f'B{row}'] = "■ 커버리지 (감사 완결성)"
        ws[f'B{row}'].font = self.FONT_SECTION
        row += 1
        try:
            unresolved = self.db.get_unresolved_assets()
        except Exception:
            unresolved = []
        if not unresolved:
            ws[f'B{row}'] = "등록된 전체 자산에 대해 KISA 판정이 모두 존재합니다 (미해결 자산 없음)."
            row += 1
        else:
            ws[f'B{row}'] = f"미해결 자산 {len(unresolved)}건 - 발견됐으나 KISA 판정 코드가 하나도 없습니다."
            row += 1
            for ip, hostname in unresolved:
                ws[f'C{row}'] = ip
                ws[f'D{row}'] = hostname or "-"
                self._style_range(ws, f'C{row}:D{row}', border=self.BORDER_ALL,
                                   alignment=Alignment(horizontal='left', indent=1))
                row += 1

    # ------------------------------------------------------------------
    # 시트: 2.점검대상
    # ------------------------------------------------------------------
    def _create_scope_sheet(self, wb, asset_data):
        ws = wb.create_sheet("2.점검대상")
        headers = [
            ("A", "No", 5), ("B", "IP 주소", 15), ("C", "Hostname", 20), ("D", "OS", 16),
            ("E", "용도", 20), ("F", "부서", 14), ("G", "담당자", 12),
            ("H", "C", 6), ("I", "I", 6), ("J", "A", 6), ("K", "등급", 8),
        ]
        for col, title, width in headers:
            cell = ws[f'{col}1']
            cell.value = title
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        for idx, (asset_id, ip, hostname, os_type, purpose, dept, owner, c, i, a) in enumerate(asset_data, 1):
            grade = DBConnector.compute_asset_grade(c, i, a)
            vals = [idx, ip, hostname or "-", os_type or "-", purpose or "-", dept or "-", owner or "-",
                    c if c is not None else "-", i if i is not None else "-", a if a is not None else "-", grade]
            r_idx = idx + 1
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = self.clean_text(val)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # ------------------------------------------------------------------
    # 시트: {카테고리}.요약통계
    # ------------------------------------------------------------------
    def _create_category_sheet(self, wb, cat_name, rows):
        """[2026-09 - "시트가 너무 많다" 피드백 반영] 예전엔 카테고리 1개당
        {cat}.요약통계(항목코드별) + {cat}.보안수준통계(분류 롤업) 2개 시트로
        나뉘어 있었다 - 서로 다른 데이터가 아니라 같은 카테고리를 보는 두 해상도
        (분류 롤업 vs 항목코드별 상세)일 뿐이라 시트 하나로 합쳤다(5개 카테고리
        기준 시트 10개 -> 5개). 롤업 표 옆에는 "그래프가 너무 단조롭다" 피드백에
        맞춰 분류별 보안수준(%) 막대 차트를 바로 붙인다."""
        sheet_name = self._unique_sheet_name(f"{cat_name}.현황")
        ws = wb.create_sheet(sheet_name)

        by_subcat = {}
        subcat_order = []
        by_code = {}
        code_order = []
        for r in rows:
            rule = self._rule_for(r["code"])
            subcat = rule.get("category", "기타") if rule else "기타"
            importance = rule.get("importance", "중") if rule else "중"
            status = self._final_status(r["status"], r["waiver_status"])

            if subcat not in by_subcat:
                by_subcat[subcat] = []
                subcat_order.append(subcat)
            by_subcat[subcat].append((importance, status))

            code = r["code"]
            if code not in by_code:
                by_code[code] = []
                code_order.append(code)
            by_code[code].append(r)

        # --- ① 분류별 보안수준 통계 (롤업) ---
        ws['A1'] = f"■ {cat_name} 분류별 보안수준 통계"
        ws['A1'].font = self.FONT_SECTION
        headers1 = [
            ("A", "분류", 16), ("B", "전체항목", 10), ("C", "점검항목", 10),
            ("D", "양호", 8), ("E", "취약", 8), ("F", "부분만족", 10), ("G", "해당없음", 10),
            ("H", "전체점수", 10), ("I", "취약점수", 10), ("J", "보안수준", 10),
        ]
        header_row1 = 2
        for col, title, width in headers1:
            cell = ws[f'{col}{header_row1}']
            cell.value = title
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col].width = width
        ws.row_dimensions[header_row1].height = 22

        r_idx = header_row1 + 1
        all_items = []
        subcat_first_row = r_idx
        for subcat in subcat_order:
            items = by_subcat[subcat]
            all_items.extend(items)
            stats = self._score_group(items)
            self._write_security_row(ws, r_idx, subcat, stats, bold=False)
            r_idx += 1
        subcat_last_row = r_idx - 1
        total_stats = self._score_group(all_items)
        self._write_security_row(ws, r_idx, "전체", total_stats, bold=True)
        r_idx += 1

        # [그래프 추가] 분류별 보안수준(%) 막대 차트 - "전체" 합계 행은 다른 분류들과
        # 스케일이 안 맞는 합산 행이라 차트에는 넣지 않고 개별 분류만 비교한다.
        if subcat_last_row >= subcat_first_row:
            self._add_security_level_chart(
                ws, f"{cat_name} 분류별 보안수준", f"L{header_row1}",
                header_row=header_row1, first_row=subcat_first_row, last_row=subcat_last_row,
                value_col=10, cat_col=1,
            )

        # [2026-09 - "한눈에 보기 편하게" 위치 조정] 차트가 표 "옆"(L열)에 떠 있고
        # 다음 섹션은 A열부터 시작하므로 열이 겹치지 않는다 - 차트 높이만큼
        # 세로 공간을 비워둘 필요가 없어서(예전엔 15행씩 빈 공간이 생겼음) 최소
        # 여백만 남긴다.
        r_idx += 2

        # --- ② 항목코드별 상세 통계 ---
        ws[f'A{r_idx}'] = f"■ {cat_name} 항목코드별 상세 통계"
        ws[f'A{r_idx}'].font = self.FONT_SECTION
        r_idx += 1
        headers2 = [
            ("A", "분류", 14), ("B", "항목코드", 12), ("C", "항목명", 40),
            ("D", "중요도", 8), ("E", "중요도값", 10),
            ("F", "전체", 8), ("G", "점검", 8), ("H", "양호", 8),
            ("I", "취약", 8), ("J", "부분만족", 10), ("K", "해당없음", 10),
        ]
        for col, title, width in headers2:
            cell = ws[f'{col}{r_idx}']
            cell.value = title
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, width)
        ws.row_dimensions[r_idx].height = 22
        ws.freeze_panes = f'A{r_idx + 1}'
        r_idx += 1

        for code in code_order:
            code_rows = by_code[code]
            rule = self._rule_for(code)
            name = rule.get("name", code_rows[0]["vuln_name"]) if rule else code_rows[0]["vuln_name"]
            category = rule.get("category", "-") if rule else "-"
            importance = rule.get("importance", "중") if rule else "중"
            items = [(importance, self._final_status(r["status"], r["waiver_status"])) for r in code_rows]
            stats = self._score_group(items)

            vals = [category, code, name, importance, IMPORTANCE_WEIGHT.get(importance, 8),
                    stats["전체"], stats["점검"], stats["양호"], stats["취약"], stats["부분만족"], stats["해당없음"]]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = self.clean_text(val)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(c_idx == 3))
                if c_idx == 3:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            r_idx += 1

    def _write_security_row(self, ws, r_idx, label, stats, bold):
        level_pct = "N/A" if stats["보안수준"] is None else stats["보안수준"]
        vals = [label, stats["전체"], stats["점검"], stats["양호"], stats["취약"],
                stats["부분만족"], stats["해당없음"], stats["전체점수"], stats["취약점수"], level_pct]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = self.FONT_BOLD if bold else self.FONT_NORMAL
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c_idx == 10 and isinstance(val, float):
                cell.value = val
                cell.number_format = '0%'
            else:
                cell.value = self.clean_text(val)
        if bold:
            self._style_range(ws, f'A{r_idx}:J{r_idx}', fill=self.HEADER_FILL)

    # ------------------------------------------------------------------
    # 시트: {카테고리}.{호스트} (호스트 1대당 상세 시트)
    # ------------------------------------------------------------------
    def _create_host_sheets(self, wb, cat_name, rows):
        """[2026-09 - 호스트 기준으로 환원, 사용자 명시적 확인] 2026-08-06엔
        "호스트가 많으면 시트가 수십~수백 개가 된다"는 이유로 카테고리당 시트
        1개(호스트 블록을 이어붙이는 방식)로 합쳤었는데, 이번엔 반대로 "호스트
        1개당 시트 1개, 그 안에 점검항목 여러 개"가 자산별로 따로 전달하기
        편하다는 피드백을 받아 다시 호스트 기준으로 되돌렸다 - 호스트 수가 많은
        현장에서는 시트 수가 다시 늘어난다는 트레이드오프를 사용자에게 명시적으로
        확인받고 진행한 것."""
        by_asset = {}
        order = []
        for r in rows:
            aid = r["asset_id"]
            if aid not in by_asset:
                by_asset[aid] = []
                order.append(aid)
            by_asset[aid].append(r)

        for aid in order:
            host_rows = by_asset[aid]
            head = host_rows[0]
            host_label = head["hostname"] or head["ip_addr"]
            sheet_name = self._unique_sheet_name(f"{cat_name}.{host_label}")
            ws = wb.create_sheet(sheet_name)
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 45
            ws.column_dimensions['I'].width = 45

            self._fill_host_block(ws, 1, head, host_rows)

    def _fill_host_block(self, ws, r, head, host_rows):
        """호스트 1대분 상세 블록을 ws의 r행부터 그려 넣고, 다음 블록이 시작할 행 번호를 반환한다."""
        ws[f'A{r}'] = "▣ 자산 정보"
        ws[f'A{r}'].font = self.FONT_SECTION
        r += 2

        info_pairs = [
            ("Host Name", head["hostname"] or "-", "IP", head["ip_addr"]),
            ("O/S", head["os_type"] or "-", "용도", head["asset_purpose"] or "-"),
            ("부서", head["department"] or "-", "담당자", head["owner_name"] or "-"),
        ]
        for label1, val1, label2, val2 in info_pairs:
            for col, label, val in ((1, label1, val1), (3, label2, val2)):
                lcell = ws.cell(row=r, column=col)
                lcell.value = label
                lcell.fill = self.HEADER_FILL
                lcell.font = self.FONT_HEADER
                lcell.border = self.BORDER_ALL
                lcell.alignment = Alignment(horizontal='center', vertical='center')
                vcell = ws.cell(row=r, column=col + 1)
                vcell.value = self.clean_text(val)
                vcell.border = self.BORDER_ALL
                vcell.font = self.FONT_NORMAL
                vcell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            r += 1
        r += 1

        c, i, a = head["conf_score"], head["integ_score"], head["avail_score"]
        grade = DBConnector.compute_asset_grade(c, i, a)
        ws[f'A{r}'] = "▣ 자산평가"
        ws[f'A{r}'].font = self.FONT_SECTION
        r += 1
        for col, label in zip((1, 2, 3, 4), ("C", "I", "A", "등급")):
            cell = ws.cell(row=r, column=col)
            cell.value = label
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center')
        r += 1
        for col, val in zip((1, 2, 3, 4), (c if c is not None else "-", i if i is not None else "-",
                                            a if a is not None else "-", grade)):
            cell = ws.cell(row=r, column=col)
            cell.value = self.clean_text(val)
            cell.font = self.FONT_BOLD
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center')
        r += 2

        # 영역별 점검 결과 (이 호스트만)
        ws[f'A{r}'] = "▣ 영역별 점검 결과"
        ws[f'A{r}'].font = self.FONT_SECTION
        r += 1
        sec_header_row = r
        sec_headers = ["분류", "전체", "점검", "양호", "취약", "부분만족", "해당없음", "전체점수", "취약점수", "보안수준"]
        for col, h in enumerate(sec_headers, 1):
            cell = ws.cell(row=r, column=col)
            cell.value = h
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center')
        r += 1

        by_subcat = {}
        order = []
        detail_rows = []
        for row in host_rows:
            rule = self._rule_for(row["code"])
            subcat = rule.get("category", "기타") if rule else "기타"
            importance = rule.get("importance", "중") if rule else "중"
            name = rule.get("name", row["vuln_name"]) if rule else row["vuln_name"]
            remediation = rule.get("remediation", row["remediation"]) if rule else row["remediation"]
            status = self._final_status(row["status"], row["waiver_status"])
            if subcat not in by_subcat:
                by_subcat[subcat] = []
                order.append(subcat)
            by_subcat[subcat].append((importance, status))
            detail_rows.append((subcat, row["code"], name, importance, status, row, remediation))

        subcat_first_row = r
        all_items = []
        for subcat in order:
            items = by_subcat[subcat]
            all_items.extend(items)
            stats = self._score_group(items)
            self._write_security_row(ws, r, subcat, stats, bold=False)
            r += 1
        subcat_last_row = r - 1
        total_stats = self._score_group(all_items)
        self._write_security_row(ws, r, "전체", total_stats, bold=True)
        r += 3

        # [그래프 추가 - 2026-09 "호스트 시트에도 차트 필요" 피드백] 이 호스트의
        # 분류별 보안수준(%) 막대 차트 - "전체" 합계 행은 스케일이 안 맞아 제외.
        if subcat_last_row >= subcat_first_row:
            self._add_security_level_chart(
                ws, "분류별 보안수준", f"L{sec_header_row}",
                header_row=sec_header_row, first_row=subcat_first_row, last_row=subcat_last_row,
                value_col=10, cat_col=1,
            )
            # [2026-09 - "한눈에 보기 편하게" 위치 조정] 차트가 표 옆(L열)에 떠
            # 있고 다음 섹션은 A열부터라 열이 안 겹친다 - 세로 공간을 비워둘
            # 필요 없이 최소 여백만 남긴다.
            r += 2

        # 상세 점검 현황
        ws[f'A{r}'] = "▣ 상세 점검 현황"
        ws[f'A{r}'].font = self.FONT_SECTION
        r += 1
        detail_headers = ["분류", "코드", "항목명", "중요도", "중요도값", "결과값", "점검결과", "현황(증적)", "조치방안"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=r, column=col)
            cell.value = h
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.border = self.BORDER_ALL
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 24
        r += 1

        for subcat, code, name, importance, status, row, remediation in detail_rows:
            weight = IMPORTANCE_WEIGHT.get(importance, 8)
            if status == "취약":
                item_score = weight
            elif status == "부분만족":
                item_score = weight / 2
            elif status == "양호":
                item_score = 0
            else:
                item_score = "-"

            evidence = row["raw_output"] or row["detected_value"]
            if evidence and len(str(evidence)) > 32000:
                evidence = str(evidence)[:32000] + "\n...[증적 내용이 너무 길어 잘렸습니다. DB를 확인하세요.]"
            if self.evidence_level == "vuln_only" and status not in ("취약", "부분만족"):
                evidence = "(상세 증적은 Professional 이상 등급에서 제공됩니다)"

            if self.remediation_level == "none":
                remediation = "(조치 방안은 Professional 이상 등급에서 제공됩니다)"
            elif self.remediation_level == "partial" and row["risk_level"] not in ("Critical", "High"):
                remediation = "(중요도 상/중 항목만 제공됩니다. 전체 제공은 Enterprise 등급)"

            vals = [subcat, code, name, importance, weight, item_score, status, evidence, remediation]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.value = self.clean_text(val)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if c_idx in (1, 2, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                if c_idx == 7:
                    cell.font = self.FONT_BOLD
                    fill = self.STATUS_FILL.get(status)
                    if fill:
                        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
            r += 1

        return r

    # ------------------------------------------------------------------
    # 시트: 4.회차비교 (기존 로직 그대로 유지 - 카테고리 재구성과 무관)
    # ------------------------------------------------------------------
    def _create_diff_sheet(self, wb, round_comparison):
        """[Phase 5: Diff 리포트] 자산별 직전 회차 대비 점수 변화(개선율)를 보여준다.
        2회 이상 스캔된 자산이 하나도 없으면 generate()에서 아예 호출되지 않는다."""
        ws = wb.create_sheet("4.회차비교(개선율)")
        headers = [
            ("A", "IP 주소", 15), ("B", "호스트명", 20),
            ("C", "직전 회차 점수", 14), ("D", "최신 회차 점수", 14), ("E", "변화", 10),
            ("F", "직전 취약 건수", 14), ("G", "최신 취약 건수", 14),
            ("H", "직전 부분만족 건수", 16), ("I", "최신 부분만족 건수", 16),
        ]
        for col, title, width in headers:
            cell = ws[f'{col}1']
            cell.value = title
            cell.fill = self.HEADER_FILL
            cell.font = self.FONT_HEADER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_ALL
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 25

        for idx, row in enumerate(round_comparison, 1):
            improvement = row["improvement"]
            vals = [
                row["ip"], row["hostname"], row["prev_score"], row["current_score"],
                ("+" if improvement > 0 else "") + str(improvement),
                row["prev_vuln_total"], row["current_vuln_total"],
                row["prev_partial_total"], row["current_partial_total"],
            ]
            r_idx = idx + 1
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = self.clean_text(val)
                cell.font = self.FONT_NORMAL
                cell.border = self.BORDER_ALL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if c_idx == 5:
                    cell.font = self.FONT_BOLD
                    if improvement > 0:
                        cell.fill = PatternFill(start_color=self.COLOR_GOOD, fill_type='solid')
                    elif improvement < 0:
                        cell.fill = PatternFill(start_color=self.COLOR_CRITICAL, fill_type='solid')
